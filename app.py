# -*- coding: utf-8 -*-
from flask import Flask, render_template_string, request, send_file, redirect, url_for
from werkzeug.middleware.proxy_fix import ProxyFix
import io
import pandas as pd
import re
from datetime import datetime

app = Flask(__name__)

# Greek-style number formatter and Jinja filter
def format_gr(value, decimals: int = 2):
    """Format numbers using Greek style: thousands '.' and decimal ','.
    Always shows exactly `decimals` decimal places. None -> ''
    """
    try:
        if value is None:
            return ''
        num = float(value)
    except Exception:
        return ''
    s = f"{num:,.{decimals}f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return s

app.jinja_env.filters['gr'] = format_gr

# In-memory storage for Production Plan grid (simple, non-persistent)
PLAN_DATA = {
    'cells': {}  # key: (line:str, date_iso:str) -> {'code': str, 'qty': float|str}
}

# Load receipes.xlsx and keep only active rows
try:
    df = pd.read_excel("receipes.xlsx")
    if "Ανενεργό" in df.columns:
        df = df[df["Ανενεργό"] == 0]
except Exception:
    df = pd.DataFrame()

# Load sales.xlsx if present
try:
    sales_df = pd.read_excel("sales.xlsx")
except Exception:
    sales_df = pd.DataFrame()

# Load stock.xlsx if present
try:
    stock_df = pd.read_excel("stock.xlsx")
except Exception:
    stock_df = pd.DataFrame()

# Load procure.xlsx if present (open purchase orders)
try:
    procure_df = pd.read_excel("procure.xlsx")
except Exception:
    procure_df = pd.DataFrame()


def _find_col(df: pd.DataFrame, candidates):
    """Find a column by exact name, else by case-insensitive contains match."""
    for c in candidates:
        if c in df.columns:
            return c
    lower = {str(c).lower(): c for c in df.columns}
    for cand in candidates:
        cl = str(cand).lower()
        for k, orig in lower.items():
            if cl in k:
                return orig
    return None


def build_stock_index(df: pd.DataFrame):
    """Return mapping code -> {total, by_storage} excluding rows with status == 2.
    Tries to auto-detect columns for code, quantity, storage, status.
    - If code is in a free-text column like 'Είδος', uses the first token (before space).
    """
    if df is None or df.empty:
        return {}

    code_col = _find_col(df, [
        'Κωδ. Αναλ.', 'Κωδικός', 'Κωδικός Υλικού', 'Κωδ. Υλικού', 'Είδος', 'Κωδ. Υλικού/Είδους'
    ])
    qty_col = _find_col(df, [
        'Υπόλοιπο', 'Διαθέσιμο', 'Ποσότητα', 'Υπόλοιπο Ποσότητας', 'Υπολ. Ποσότητα'
    ])
    storage_col = _find_col(df, [
        'Α.Χ.', 'Αποθηκευτικός χώρος', 'Αποθηκ. Χώρος', 'Αποθ. Χώρος', 'Θέση', 'Αποθήκη', 'Χώρος'
    ])
    status_col = _find_col(df, [
        'Status', 'Στατους', 'Κατάσταση', 'Στάτους', 'Κατάσταση παρτίδας'
    ])

    if not code_col or not qty_col:
        return {}

    work = df.copy()
    if status_col and status_col in work.columns:
        try:
            work = work[work[status_col].astype(str) != '2']
        except Exception:
            pass

    index = {}
    for _, row in work.iterrows():
        code_text = str(row.get(code_col, '')).strip()
        code = code_text.split()[0] if code_text else ''
        if not code:
            continue
        try:
            qty = float(row.get(qty_col, 0) or 0)
        except Exception:
            qty = 0
        loc = str(row.get(storage_col, '')).strip() if storage_col in work.columns else ''
        item = index.setdefault(code, {'total': 0.0, 'by_storage': {}})
        item['total'] += qty
        if loc:
            item['by_storage'][loc] = item['by_storage'].get(loc, 0.0) + qty
    return index


def build_procure_index(df: pd.DataFrame):
    """Return mapping code -> expected incoming quantity (sum of open qty).
    Tries to auto-detect columns for code and open quantity.
    """
    if df is None or df.empty:
        return {}

    code_col = _find_col(df, [
        'Κωδικός Είδους', 'Κωδ. Υλικού', 'Κωδικός', 'Είδος', 'Κωδ. Υλικού/Είδους'
    ])
    open_qty_col = _find_col(df, [
        'Ανοιχτή ποσότητα', 'Υπόλοιπο', 'Υπόλοιπο Ποσότητας', 'Υπολ. Ποσότητα'
    ])
    if not code_col or not open_qty_col:
        return {}

    index = {}
    for _, row in df.iterrows():
        code_text = str(row.get(code_col, '')).strip()
        token = code_text.split()[0] if code_text else ''
        code = _base_code(token)
        if not code:
            continue
        try:
            qty = float(row.get(open_qty_col, 0) or 0)
        except Exception:
            qty = 0
        if qty <= 0:
            continue
        index[code] = index.get(code, 0.0) + qty
    return index


def build_procure_by_date(df: pd.DataFrame):
    """Return mapping code -> {date_str: qty} using 'Ημ/νία 5' as delivery date.
    If 'Ημ/νία 5' is missing or unparsable, fallback to 'Ημ/νία παράδοσης'.
    """
    if df is None or df.empty:
        return {}

    code_col = _find_col(df, [
        'Κωδικός Είδους', 'Κωδ. Υλικού', 'Κωδικός', 'Είδος', 'Κωδ. Υλικού/Είδους'
    ])
    open_qty_col = _find_col(df, [
        'Ανοιχτή ποσότητα', 'Υπόλοιπο', 'Υπόλοιπο Ποσότητας', 'Υπολ. Ποσότητα'
    ])
    date5_col = _find_col(df, [
        'Ημ/νία 5', 'Ημερομηνία 5'
    ])
    delivery_col = _find_col(df, [
        'Ημ/νία παράδοσης', 'Ημερομηνία παράδοσης', 'Παράδοση'
    ])
    
    if not code_col or not open_qty_col or (not date5_col and not delivery_col):
        return {}

    # Prepare columns to work with
    cols_to_use = [code_col, open_qty_col]
    if date5_col:
        cols_to_use.append(date5_col)
    if delivery_col:
        cols_to_use.append(delivery_col)
    
    work = df[cols_to_use].copy()
    # Normalize
    work[code_col] = work[code_col].astype(str).str.strip().str.split().str[0].fillna('')
    work[code_col] = work[code_col].apply(_base_code)
    work[open_qty_col] = pd.to_numeric(work[open_qty_col], errors='coerce').fillna(0.0)
    
    # Parse date columns
    if date5_col:
        work[date5_col] = pd.to_datetime(work[date5_col], errors='coerce')
    if delivery_col:
        work[delivery_col] = pd.to_datetime(work[delivery_col], errors='coerce')
    
    # Filter for valid codes and positive quantities
    work = work[(work[code_col] != '') & (work[open_qty_col] > 0)]

    index: dict[str, dict[str, float]] = {}
    for _, row in work.iterrows():
        code = row[code_col]
        qty = float(row[open_qty_col])
        
        # Try Ημ/νία 5 first, then fallback to Ημ/νία παράδοσης
        date_to_use = None
        if date5_col and pd.notna(row[date5_col]):
            date_to_use = row[date5_col]
        elif delivery_col and pd.notna(row[delivery_col]):
            date_to_use = row[delivery_col]
        
        if date_to_use is not None:
            dstr = date_to_use.date().isoformat()
            bucket = index.setdefault(code, {})
            bucket[dstr] = bucket.get(dstr, 0.0) + qty
    return index
def get_procure_missing_date5(df: pd.DataFrame):
    """Return (DataFrame, code_col, delivery_col) of purchase rows with positive open qty and missing 'Ημ/νία 5'.
    Includes delivery date column if present. Intentionally excludes the empty 'Ημ/νία 5' column from the output.
    """
    if df is None or df.empty:
        return pd.DataFrame(), None, None
    code_col = _find_col(df, ['Κωδικός Είδους','Κωδ. Υλικού','Κωδικός','Είδος','Κωδ. Υλικού/Είδους'])
    ord_col = _find_col(df, [
        'Αρ. Παραγγελίας', 'Αριθμός Παραγγελίας', 'Παραγγελία', 'Παραγγελία Αγοράς',
        'Αρ. Εγγράφου', 'Αριθμός Εγγράφου', 'Εντολή αγοράς', 'Εντολή Αγοράς'
    ])
    desc_col = _find_col(df, ['Περιγραφή','Περιγραφή Είδους','Περιγραφή Υλικού'])
    open_qty_col = _find_col(df, ['Ανοιχτή ποσότητα','Υπόλοιπο','Υπόλοιπο Ποσότητας','Υπολ. Ποσότητα'])
    date5_col = _find_col(df, ['Ημ/νία 5','Ημερομηνία 5'])
    delivery_col = _find_col(df, ['Ημ/νία παράδοσης','Ημερομηνία παράδοσης','Παράδοση'])
    sup_col = _find_col(df, ['Προμηθευτής','Προμηθευτής/Όνομα','Προμηθευτής-Όνομα'])
    if not code_col or not open_qty_col:
        return pd.DataFrame(), code_col, delivery_col
    work = df.copy()
    work['__open__'] = pd.to_numeric(work[open_qty_col], errors='coerce').fillna(0.0)
    if date5_col:
        work['__date5__'] = pd.to_datetime(work[date5_col], errors='coerce')
        mask_missing = work['__date5__'].isna()
    else:
        mask_missing = pd.Series([True]*len(work))
    work = work[(work['__open__'] > 0) & mask_missing]
    # Do not include date5_col (it's missing by definition in this report). Include order number if present.
    cols = [c for c in [ord_col, code_col, desc_col, open_qty_col, delivery_col, sup_col] if c and c in work.columns]
    return work[cols].copy(), code_col, delivery_col

PROCURE_MISSING_TEMPLATE = r"""
<!doctype html>
<html>
<head>
    <meta charset="utf-8">
    <title>Αγορές χωρίς Ημ/νία 5</title>
    <style>
        body { background:#fff; color:#000; font-family:sans-serif; margin:0; }
        .container { padding:16px 20px; }
        table { width:100%; border-collapse: collapse; margin-top: 12px; }
        th, td { border:1px solid #000; padding:4px; text-align:left; }
        .num { text-align:right; font-variant-numeric: tabular-nums; }
        .btn { display:inline-block; padding:8px 12px; border:1px solid #000; background:#eee; color:#000; text-decoration:none; cursor:pointer; }
    </style>
        <script>
            function textOf(cell){ return (cell.innerText || cell.textContent || '').trim(); }
            function toggleHideNonPattern(){
                var tbl = document.getElementById('miss_table');
                var chk = document.getElementById('hide_non_pattern');
                if (!tbl || !chk) return;
                var hide = chk.checked;
                var codeIdx = parseInt(tbl.getAttribute('data-code-col') || '-1');
                var rows = tbl.getElementsByTagName('tr');
                for (var i=1; i<rows.length; i++){
                    var tds = rows[i].getElementsByTagName('td');
                    var code = (codeIdx>=0 && codeIdx<tds.length) ? textOf(tds[codeIdx]) : '';
                    var ok = /^\d{3}-\d{2}-\d{2}$/.test(code);
                    rows[i].style.display = (hide && !ok) ? 'none' : '';
                }
                localStorage.setItem('miss_hide_non_pattern', hide ? '1':'0');
            }
            function initHideNonPattern(){
                var chk = document.getElementById('hide_non_pattern');
                if (!chk) return;
                chk.checked = localStorage.getItem('miss_hide_non_pattern') === '1';
                toggleHideNonPattern();
            }
            // --- Sorting helpers ---
            function parseNumberGR(s){ if(s==null) return 0; s=(''+s).trim(); if(s==='') return 0; s=s.replace(/\./g,'').replace(',', '.'); var v=parseFloat(s); return isNaN(v)?0:v; }
            function parseDMY(s){ var m=(''+(s||'')).match(/^(\d{1,2})\/(\d{1,2})\/(\d{4})$/); if(!m) return 0; return new Date(parseInt(m[3],10), parseInt(m[2],10)-1, parseInt(m[1],10)).getTime(); }
            function sortMissTable(colIndex, type){
                var tbl=document.getElementById('miss_table'); if(!tbl) return;
                var rows=Array.prototype.slice.call(tbl.querySelectorAll('tr'));
                if(rows.length<2) return; var header=rows.shift();
                var currentCol=tbl.getAttribute('data-sort-col'); var currentDir=tbl.getAttribute('data-sort-dir')||'asc';
                var dir=(currentCol===String(colIndex)&&currentDir==='asc')?'desc':'asc'; var mod=dir==='asc'?1:-1;
                rows.sort(function(a,b){
                    var A=(a.getElementsByTagName('td')[colIndex]||{}).innerText||'';
                    var B=(b.getElementsByTagName('td')[colIndex]||{}).innerText||'';
                    if(type==='number'){ var na=parseNumberGR(A), nb=parseNumberGR(B); if(na<nb) return -1*mod; if(na>nb) return 1*mod; return 0; }
                    if(type==='date'){ var da=parseDMY(A), db=parseDMY(B); if(da<db) return -1*mod; if(da>db) return 1*mod; return 0; }
                    A=A.toLowerCase(); B=B.toLowerCase(); var cmp=A.localeCompare(B,'el',{numeric:true,sensitivity:'base'}); return cmp*mod;
                });
                var tbody=tbl.tBodies && tbl.tBodies[0] ? tbl.tBodies[0] : tbl;
                while (tbody.firstChild) tbody.removeChild(tbody.firstChild);
                tbody.appendChild(header); for (var i=0;i<rows.length;i++) tbody.appendChild(rows[i]);
                tbl.setAttribute('data-sort-col', String(colIndex)); tbl.setAttribute('data-sort-dir', dir);
                var ths=header.getElementsByTagName('th'); for (var j=0;j<ths.length;j++){ ths[j].removeAttribute('data-sorted'); ths[j].style.cursor='pointer'; }
                if (ths[colIndex]) ths[colIndex].setAttribute('data-sorted', dir);
            }
            function initMissSortable(){
                var tbl=document.getElementById('miss_table'); if(!tbl) return;
                var header=tbl.querySelector('tr'); if(!header) return;
                var ths=header.getElementsByTagName('th');
                for (var i=0;i<ths.length;i++){
                    (function(idx){ var th=ths[idx]; var t=th.getAttribute('data-sort')||'text'; if(t==='none') return; th.style.cursor='pointer'; th.onclick=function(){ sortMissTable(idx,t); }; })(i);
                }
            }
            window.addEventListener('load', initHideNonPattern);
            window.addEventListener('load', initMissSortable);
        </script>
</head>
<body>
    <div class="container">
        <h2>Αγορές χωρίς "Ημ/νία 5"</h2>
            <p>
                <a class="btn" href="/">Αρχική</a>
                <label style="margin-left:12px;">
                    <input type="checkbox" id="hide_non_pattern" onchange="toggleHideNonPattern()"> Κρύψε κωδικούς εκτός μορφής xxx-xx-xx
                </label>
                <a class="btn" style="margin-left:12px;" href="/procure-missing-date5?download=1">Κατέβασμα σε Excel</a>
            </p>
        {% if rows %}
            <table id="miss_table" data-code-col="{{ code_col_index }}">
            <tr>
                {% for h in headers %}
                    {% set lower = h|lower %}
                    {% if 'ημ/νία' in lower or 'ημερομην' in lower %}
                        <th data-sort="date">{{ h }}</th>
                    {% elif 'ποσ' in lower %}
                        <th data-sort="number">{{ h }}</th>
                    {% else %}
                        <th data-sort="text">{{ h }}</th>
                    {% endif %}
                {% endfor %}
            </tr>
            {% for r in rows %}
            <tr>
                {% for h in headers %}
                    {% set v = r.get(h, '') %}
                    {% if h.lower().find('ποσ') != -1 %}
                        <td class="num">{{ v|gr(2) }}</td>
                    {% else %}
                            <td>{{ v }}</td>
                    {% endif %}
                {% endfor %}
            </tr>
            {% endfor %}
        </table>
        {% else %}
            <p>Δεν βρέθηκαν εγγραφές.</p>
        {% endif %}
    </div>
</body>
</html>
"""

# INACTIVE ROUTE - uncomment @app.route to re-enable
# @app.route('/procure-missing-date5')
def procure_missing_date5():
    dfmiss, code_col, delivery_col = get_procure_missing_date5(procure_df)
    # Format any date-like columns to dd/mm/yyyy (delivery, date5, etc.)
    if not dfmiss.empty:
        for col in list(dfmiss.columns):
            cl = str(col).lower()
            if 'ημ/νία' in cl or 'ημερομην' in cl:
                try:
                    s = pd.to_datetime(dfmiss[col], errors='coerce')
                    dfmiss[col] = s.dt.strftime('%d/%m/%Y').fillna('')
                except Exception:
                    pass
        # If download requested, export as Excel
        if request.args.get('download') == '1':
            out_df = dfmiss.copy()
            # Try to order columns: Order No, Code, Desc, Open Qty, Delivery, Supplier (if present)
            cols_pref = []
            # Locate likely columns by matching substrings in lowercase
            def _find_by(subs):
                for c in out_df.columns:
                    cl = str(c).lower()
                    if any(s in cl for s in subs):
                        return c
                return None
            ord_c = _find_by(['παραγγελ', 'εγγράφ', 'εντολή'])
            code_c = code_col
            desc_c = _find_by(['περιγραφ'])
            open_c = _find_by(['ανοιχτή ποσ', 'υπόλοιπο'])
            deliv_c = delivery_col
            sup_c = _find_by(['προμηθευ'])
            for c in [ord_c, code_c, desc_c, open_c, deliv_c, sup_c]:
                if c and c in out_df.columns and c not in cols_pref:
                    cols_pref.append(c)
            # Add remaining columns
            cols_pref += [c for c in out_df.columns if c not in cols_pref]
            out_df = out_df[cols_pref]

            buf = io.BytesIO()
            try:
                with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
                    out_df.to_excel(writer, index=False, sheet_name='MissingDate5')
                pass
            except Exception:
                with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                    out_df.to_excel(writer, index=False, sheet_name='MissingDate5')
            buf.seek(0)
            return send_file(
                buf,
                as_attachment=True,
                download_name=f"missing_date5_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    headers = list(dfmiss.columns) if not dfmiss.empty else []
    rows = dfmiss.to_dict(orient='records') if not dfmiss.empty else []
    code_col_index = headers.index(code_col) if (code_col and code_col in headers) else -1
    return render_template_string(PROCURE_MISSING_TEMPLATE, headers=headers, rows=rows, code_col_index=code_col_index)


def build_stock_lots(df: pd.DataFrame):
    """Build per-lot stock records with storage and expiry info.
    Returns mapping: code -> [ {storage, expiry(pd.Timestamp or None), qty(float)} ]
    Excludes rows with status == 2. Tries to auto-detect columns.
    """
    if df is None or df.empty:
        return {}

    code_col = _find_col(df, [
        'Κωδ. Αναλ.', 'Κωδικός', 'Κωδικός Υλικού', 'Κωδ. Υλικού', 'Είδος', 'Κωδ. Υλικού/Είδους'
    ])
    qty_col = _find_col(df, [
        'Υπόλοιπο', 'Διαθέσιμο', 'Ποσότητα', 'Υπόλοιπο Ποσότητας', 'Υπολ. Ποσότητα'
    ])
    storage_col = _find_col(df, [
        'Α.Χ.', 'Αποθηκευτικός χώρος', 'Αποθηκ. Χώρος', 'Αποθ. Χώρος', 'Θέση', 'Αποθήκη', 'Χώρος'
    ])
    status_col = _find_col(df, [
        'Status', 'Στατους', 'Κατάσταση', 'Στάτους', 'Κατάσταση παρτίδας'
    ])
    expiry_col = _find_col(df, [
        'Ημ/νία Λήξης', 'Ημ/νία λήξης', 'Ημερομηνία Λήξης', 'Ημερομηνία λήξης', 'Λήξη',
        'Ημ/νία λήξης παρτίδας', 'Ημ/νία λήξης παρτιδας', 'Expiration Date'
    ])
    batch_col = _find_col(df, [
        'Παρτίδα', 'Αρ. Παρτίδας', 'Αριθμός Παρτίδας', 'Lot', 'Batch', 'Batch No', 'Lot No'
    ])

    if not code_col or not qty_col:
        return {}

    work = df.copy()
    if status_col and status_col in work.columns:
        try:
            work = work[work[status_col].astype(str) != '2']
        except Exception:
            pass

    lots = {}
    for _, row in work.iterrows():
        code_text = str(row.get(code_col, '')).strip()
        code = code_text.split()[0] if code_text else ''
        if not code:
            continue
        try:
            qty = float(row.get(qty_col, 0) or 0)
        except Exception:
            qty = 0.0
        if qty <= 0:
            continue
        storage = str(row.get(storage_col, '')).strip() if storage_col in work.columns else ''
        expiry = None
        if expiry_col and expiry_col in work.columns:
            try:
                expiry = pd.to_datetime(row.get(expiry_col), errors='coerce')
            except Exception:
                expiry = None
        batch = str(row.get(batch_col, '')).strip() if batch_col and batch_col in work.columns else ''
        lots.setdefault(code, []).append({'storage': storage, 'expiry': expiry if pd.notna(expiry) else None, 'qty': qty, 'batch': batch})
    return lots


def get_storage_list(df: pd.DataFrame):
    """Return sorted unique list of storage locations (Α.Χ.) excluding rows with status==2."""
    if df is None or df.empty:
        return []
    storage_col = _find_col(df, [
        'Α.Χ.', 'Αποθηκευτικός χώρος', 'Αποθηκ. Χώρος', 'Αποθ. Χώρος', 'Θέση', 'Αποθήκη', 'Χώρος'
    ])
    status_col = _find_col(df, [
        'Status', 'Στατους', 'Κατάσταση', 'Στάτους', 'Κατάσταση παρτίδας'
    ])
    if not storage_col:
        return []
    work = df.copy()
    if status_col and status_col in work.columns:
        try:
            work = work[work[status_col].astype(str) != '2']
        except Exception:
            pass
    vals = sorted({str(v).strip() for v in work[storage_col].dropna().unique() if str(v).strip()})
    return vals


def get_revision_number(code):
    match = re.search(r"/(\d+)$", str(code))
    return int(match.group(1)) if match else 0

def _base_code(text: str) -> str:
    """Return the base code without revision or trailing description.
    Examples: '121-00-01/5' -> '121-00-01', '121-00-01 Κάτι' -> '121-00-01'.
    """
    s = str(text or '').strip()
    if not s:
        return ''
    s = s.split()[0]
    if '/' in s:
        s = s.split('/')[0]
    return s

_SEMI_REGEX = re.compile(r'^\d{2}2-\d{2}-\d{2}$')

def _is_semi_finished(code: str) -> bool:
    return bool(_SEMI_REGEX.match(_base_code(code)))

def _select_latest_recipe(product_code: str):
    """Select rows of the latest recipe for the given product base code.
    Prefers column 'Κωδικός Είδους Συνταγής'. Returns (rows, base_qty, description, max_rev)
    or (None, 0, '', 0) if not found.
    """
    if df.empty:
        return None, 0, '', 0
    base = _base_code(product_code)
    work = df.copy()
    target_col = 'Κωδικός Είδους Συνταγής' if 'Κωδικός Είδους Συνταγής' in work.columns else None
    if target_col:
        try:
            work['__base__'] = work[target_col].astype(str).str.replace(r'/(\d+)$', '', regex=True)
        except Exception:
            work['__base__'] = work[target_col].astype(str)
        subset = work[work['__base__'] == base].copy()
        if subset.empty:
            return None, 0, '', 0
        subset['rev_num'] = subset[target_col].apply(get_revision_number)
    else:
        # Fallback on 'Κωδικός' exact match of base
        if 'Κωδικός' not in work.columns:
            return None, 0, '', 0
        subset = work[work['Κωδικός'].astype(str).str.split('/').str[0] == base].copy()
        if subset.empty:
            return None, 0, '', 0
        subset['rev_num'] = subset['Κωδικός'].apply(get_revision_number)

    max_rev = subset['rev_num'].max()
    selected = subset[subset['rev_num'] == max_rev].copy()
    try:
        base_qty = float(selected['Ποσότητα Παραγόμενου'].iloc[0])
    except Exception:
        base_qty = 0
    description = str(selected['Περιγραφή'].iloc[0]) if 'Περιγραφή' in selected.columns and not selected.empty else ''
    return selected, base_qty, description, int(max_rev) if pd.notna(max_rev) else 0

def _per_unit_materials(product_code: str, cache: dict, visited: set):
    """Return aggregated leaf materials for producing 1 unit of product_code.
    Keys are (code, desc, unit) and values are quantities per 1 unit.
    """
    base = _base_code(product_code)
    if base in cache:
        return cache[base]
    if base in visited:
        return {}  # cycle guard
    visited.add(base)

    selected, base_qty, _, _ = _select_latest_recipe(base)
    if selected is None or base_qty == 0:
        visited.discard(base)
        return {}

    agg = {}
    for _, row in selected.iterrows():
        comp_code = str(row.get('Κωδ. Αναλ.', '')).strip()
        comp_desc = str(row.get('Περιγραφή Αναλούμενου', '')).strip()
        comp_unit = str(row.get('Μονάδα μέτρησης', '')).strip()
        try:
            per_unit = float(row.get('Ποσότητα', 0) or 0) / base_qty
        except Exception:
            per_unit = 0
        if per_unit == 0 or not comp_code:
            continue
        if _is_semi_finished(comp_code):
            # If the semi-finished has a recipe, expand; else treat as leaf
            child_rows, child_base_qty, _, _ = _select_latest_recipe(comp_code)
            if child_rows is not None and child_base_qty != 0:
                child_map = _per_unit_materials(comp_code, cache, visited)
                for k, v in child_map.items():
                    agg[k] = agg.get(k, 0.0) + v * per_unit
                continue
        # treat as leaf
        key = (comp_code, comp_desc, comp_unit)
        agg[key] = agg.get(key, 0.0) + per_unit

    visited.discard(base)
    cache[base] = agg
    return agg


def get_materials_multiple(items):
    """Aggregate materials for a list of {product_code, qty} dicts with recursive expansion of semi-finished items.
    Returns: (materials_list, details_list)
    materials_list: [{code, desc, qty, unit}]
    details_list: [{product_code, description, max_rev, not_found}]
    """
    all_materials = {}
    details = []
    cache = {}

    if df.empty:
        return [], [{"product_code": it.get("product_code", ""), "description": "", "max_rev": "", "not_found": True} for it in items]

    for item in items:
        product_code = str(item.get('product_code', '')).strip()
        try:
            qty = float(item.get('qty', 0) or 0)
        except Exception:
            qty = 0
        if not product_code or qty == 0:
            details.append({'product_code': product_code, 'description': '', 'max_rev': '', 'not_found': True})
            continue

        selected, base_qty, description, max_rev = _select_latest_recipe(product_code)
        if selected is None or base_qty == 0:
            details.append({'product_code': product_code, 'description': '', 'max_rev': '', 'not_found': True})
            continue

        details.append({'product_code': _base_code(product_code), 'description': description, 'max_rev': max_rev, 'not_found': False})

        # Compute per-unit leaf materials and scale by requested quantity
        per_unit_map = _per_unit_materials(product_code, cache, set())
        for k, per1 in per_unit_map.items():
            needed = per1 * qty
            all_materials[k] = all_materials.get(k, 0.0) + needed

    materials = [{"code": k[0], "desc": k[1], "qty": round(v, 3), "unit": k[2]} for k, v in all_materials.items()]
    return materials, details

# ------------------------------------------------------------
# ΝΕΟΣ ΜΗΧΑΝΙΣΜΟΣ: Συγκέντρωση ημιέτοιμων (semi-finished) σε όλα τα βάθη
# ------------------------------------------------------------
def _per_unit_semis(product_code: str, cache: dict, visited: set):
    """Επιστρέφει dict με (code, desc, unit) -> ποσότητα ημιέτοιμου ανά 1 τεμ. του τελικού προϊόντος.
    Αναδρομικά διατρέχει τη συνταγή και όταν συναντά ημιέτοιμο (xx2-xx-xx) το καταγράφει,
    μετά συνεχίζει μέσα στη δική του συνταγή για τυχόν nested ημιέτοιμα.
    """
    base = _base_code(product_code)
    if base in cache:
        return cache[base]
    if base in visited:
        return {}
    visited.add(base)

    selected, base_qty, _, _ = _select_latest_recipe(base)
    if selected is None or base_qty == 0:
        visited.discard(base)
        cache[base] = {}
        return {}

    agg = {}
    for _, row in selected.iterrows():
        comp_code = str(row.get('Κωδ. Αναλ.', '')).strip()
        comp_desc = str(row.get('Περιγραφή Αναλούμενου', '')).strip()
        comp_unit = str(row.get('Μονάδα μέτρησης', '')).strip()
        if not comp_code:
            continue
        try:
            per_unit = float(row.get('Ποσότητα', 0) or 0) / base_qty
        except Exception:
            per_unit = 0
        if per_unit == 0:
            continue

        # Εάν υπάρχει συνταγή για το component, κατέβα βαθύτερα για να συλλέξεις ημιέτοιμα
        child_rows, child_base_qty, _, _ = _select_latest_recipe(comp_code)
        if child_rows is not None and child_base_qty != 0:
            nested_map = _per_unit_semis(comp_code, cache, visited)
            for nk, nv in nested_map.items():
                agg[nk] = agg.get(nk, 0.0) + nv * per_unit

        # Αν είναι ημιέτοιμο, πρόσθεσέ το και ως άμεση απαίτηση
        if _is_semi_finished(comp_code):
            key = ( _base_code(comp_code), comp_desc, comp_unit )
            agg[key] = agg.get(key, 0.0) + per_unit

    visited.discard(base)
    cache[base] = agg
    return agg

def get_semis_multiple(items):
    """Συγκεντρώνει ημιέτοιμα για λίστα από {product_code, qty}.
    Επιστρέφει (semis_list, details_list) όπου semis_list = [{code, desc, qty, unit}], qty = συνολική απαίτηση.
    """
    all_semis = {}
    details = []
    cache = {}
    if df.empty:
        return [], [{"product_code": it.get('product_code',''), "description": '', "max_rev": '', "not_found": True} for it in items]

    for item in items:
        product_code = str(item.get('product_code', '')).strip()
        try:
            qty = float(item.get('qty', 0) or 0)
        except Exception:
            qty = 0
        if not product_code or qty == 0:
            details.append({'product_code': product_code, 'description': '', 'max_rev': '', 'not_found': True})
            continue

        selected, base_qty, description, max_rev = _select_latest_recipe(product_code)
        if selected is None or base_qty == 0:
            details.append({'product_code': product_code, 'description': '', 'max_rev': '', 'not_found': True})
            continue
        details.append({'product_code': _base_code(product_code), 'description': description, 'max_rev': max_rev, 'not_found': False})

        per_unit_semis = _per_unit_semis(product_code, cache, set())
        for k, per1 in per_unit_semis.items():
            needed = per1 * qty
            all_semis[k] = all_semis.get(k, 0.0) + needed

    semis = [{"code": k[0], "desc": k[1], "qty": round(v, 3), "unit": k[2]} for k, v in all_semis.items()]
    return semis, details


def build_reverse_usage_index():
    """Build a reverse index: leaf material base code -> set of product base codes that consume it.
    Also return a map for product descriptions.
    """
    if df.empty:
        return {}, {}
    products = set()
    work = df.copy()
    target_col = 'Κωδικός Είδους Συνταγής' if 'Κωδικός Είδους Συνταγής' in work.columns else None
    if target_col:
        try:
            work['__base__'] = work[target_col].astype(str).str.replace(r'/\d+$', '', regex=True)
        except Exception:
            work['__base__'] = work[target_col].astype(str).str.split('/').str[0]
        products.update(work['__base__'].dropna().unique().tolist())
    elif 'Κωδικός' in work.columns:
        products.update(work['Κωδικός'].astype(str).str.split('/').str[0].dropna().unique().tolist())
    else:
        return {}, {}

    reverse_map = {}
    prod_desc = {}
    cache = {}
    for p in sorted(products):
        try:
            _, _, desc, _ = _select_latest_recipe(p)
            prod_desc[p] = desc
            per_unit = _per_unit_materials(p, cache, set())
            for (leaf_code, _, _), _qty in per_unit.items():
                leaf_base = _base_code(leaf_code)
                s = reverse_map.setdefault(leaf_base, set())
                s.add(p)
        except Exception:
            continue
    return reverse_map, prod_desc


TEMPLATE = r"""
<!doctype html>
<html>
<head>
    <title>Υπολογισμός Υλικών Παραγωγής</title>
    <style>
    body { background: #fff; color: #000; font-family: sans-serif; margin: 0; }
    .container { width: 100%; max-width: none; margin: 0; padding: 16px 20px; border: 1px solid #000; box-sizing: border-box; }
        input, button, textarea { font-size: 1em; margin: 5px 0; }
        table { width: 100%; border-collapse: collapse; margin-top: 12px; }
        th, td { border: 1px solid #000; padding: 4px; text-align: left; }
    .num { text-align: right; font-variant-numeric: tabular-nums; }
        .error { color: red; }
    .danger { color: #b00000; font-weight: 600; }
    .page-header { display: flex; align-items: center; justify-content: space-between; margin-bottom: 12px; }
    .page-header h2 { margin: 0; }
    /* Manual modal */
    #manual_modal_overlay {
        display: none;
        position: fixed;
        inset: 0;
        background: rgba(0,0,0,0.45);
        z-index: 9000;
        align-items: center;
        justify-content: center;
    }
    #manual_modal_overlay.active { display: flex; }
    #manual_modal_box {
        background: #fff;
        border: 1px solid #000;
        width: 60vw;
        max-width: 700px;
        max-height: 80vh;
        display: flex;
        flex-direction: column;
    }
    /* Hide storage column when the table has the hide-storage class */
    #materials_table.hide-storage th.col-storage,
    #materials_table.hide-storage td.col-storage { display: none; }
    #materials_table.hide-expected-bydate th.col-expected-bydate,
    #materials_table.hide-expected-bydate td.col-expected-bydate { display: none; }
    #materials_table.hide-usedin th.col-usedin,
    #materials_table.hide-usedin td.col-usedin { display: none; }
    #materials_table.hide-bring708 th.col-bring708,
    #materials_table.hide-bring708 td.col-bring708 { display: none; }
    th.sortable { cursor: pointer; user-select: none; }
    th.sortable::after { content: ""; padding-left: 6px; font-size: 0.9em; color:#333; }
    th.sortable[data-sorted="asc"]::after { content: "▲"; }
    th.sortable[data-sorted="desc"]::after { content: "▼"; }
    /* Orders modal */
    #orders_modal_overlay {
        display: none;
        position: fixed;
        inset: 0;
        background: rgba(0,0,0,0.45);
        z-index: 9000;
        align-items: center;
        justify-content: center;
    }
    #orders_modal_overlay.active {
        display: flex;
    }
    #orders_modal_box {
        background: #fff;
        border: 1px solid #000;
        width: 92vw;
        max-width: 1200px;
        max-height: 88vh;
        display: flex;
        flex-direction: column;
        position: relative;
    }
    /* Filters modal */
    #filters_modal_overlay {
        display: none;
        position: fixed;
        inset: 0;
        background: rgba(0,0,0,0.45);
        z-index: 9000;
        align-items: center;
        justify-content: center;
    }
    #filters_modal_overlay.active {
        display: flex;
    }
    #filters_modal_box {
        background: #fff;
        border: 1px solid #000;
        width: 80vw;
        max-width: 900px;
        max-height: 80vh;
        display: flex;
        flex-direction: column;
        position: relative;
    }
    /* Shared modal parts */
    #orders_modal_box .modal-header,
    #filters_modal_box .modal-header,
    #manual_modal_box .modal-header {
        display: flex;
        align-items: center;
        gap: 8px;
        padding: 12px 16px;
        border-bottom: 1px solid #ccc;
        flex-shrink: 0;
    }
    #orders_modal_box .modal-header h3,
    #filters_modal_box .modal-header h3,
    #manual_modal_box .modal-header h3 { margin: 0; font-size: 1em; margin-right: auto; }
    #orders_modal_box .modal-body,
    #filters_modal_box .modal-body,
    #manual_modal_box .modal-body {
        overflow-y: auto;
        padding: 12px 16px;
        flex: 1;
    }
    #orders_modal_box .modal-footer,
    #filters_modal_box .modal-footer,
    #manual_modal_box .modal-footer {
        padding: 8px 16px;
        text-align: right;
        border-top: 1px solid #ccc;
        flex-shrink: 0;
    }
    #orders_modal_close_btn,
    #filters_modal_close_btn,
    #manual_modal_close_btn {
        cursor: pointer;
        font-size: 1.4em;
        border: none;
        background: none;
        line-height: 1;
        padding: 0 4px;
        flex-shrink: 0;
    }
    .ax-filter-grid {
        display: flex;
        flex-wrap: wrap;
        gap: 4px 16px;
        padding: 4px 0;
    }
    .ax-filter-grid label {
        min-width: 120px;
        display: flex;
        align-items: center;
        gap: 4px;
        cursor: pointer;
    }
    </style>
    <script>
        window.onload = function() {
            updateToggleSelectAllLabel();
        }
        function filterSales() {
            var input = document.getElementById('search').value.toLowerCase();
            var table = document.getElementById('sales_table');
            var trs = table.getElementsByTagName('tr');
            for (var i = 1; i < trs.length; i++) {
                var tds = trs[i].getElementsByTagName('td');
                var show = false;
                for (var j = 1; j < tds.length; j++) {
                    if (tds[j].innerText.toLowerCase().indexOf(input) !== -1) { show = true; break; }
                }
                trs[i].style.display = show ? '' : 'none';
            }
        }
        function toggleOrders() {
            var overlay = document.getElementById('orders_modal_overlay');
            if (!overlay) return;
            overlay.classList.add('active');
            document.body.style.overflow = 'hidden';
        }
        function closeOrdersModal() {
            var overlay = document.getElementById('orders_modal_overlay');
            if (overlay) overlay.classList.remove('active');
            document.body.style.overflow = '';
        }
        function openManualModal() {
            document.getElementById('manual_modal_overlay').classList.add('active');
            document.body.style.overflow = 'hidden';
        }
        function closeManualModal() {
            document.getElementById('manual_modal_overlay').classList.remove('active');
            document.body.style.overflow = '';
        }
        function openFiltersModal() {
            var overlay = document.getElementById('filters_modal_overlay');
            if (!overlay) return;
            overlay.classList.add('active');
            document.body.style.overflow = 'hidden';
        }
        function closeFiltersModal() {
            var overlay = document.getElementById('filters_modal_overlay');
            if (overlay) overlay.classList.remove('active');
            document.body.style.overflow = '';
        }
        function selectAllFilters() {
            document.getElementsByName('ax_filter').forEach(function(cb){ cb.checked = true; });
        }
        function clearAllFilters() {
            document.getElementsByName('ax_filter').forEach(function(cb){ cb.checked = false; });
        }
        // Close modals on overlay click or Escape
        document.addEventListener('click', function(e) {
            var o = document.getElementById('orders_modal_overlay');
            if (o && e.target === o) { closeOrdersModal(); }
            var f = document.getElementById('filters_modal_overlay');
            if (f && e.target === f) { closeFiltersModal(); }
            var m = document.getElementById('manual_modal_overlay');
            if (m && e.target === m) { closeManualModal(); }
        });
        document.addEventListener('keydown', function(e) {
            if (e.key === 'Escape') { closeOrdersModal(); closeFiltersModal(); closeManualModal(); }
        });
        function clearSalesSelections() {
            var boxes = document.getElementsByName('sales_select');
            for (var i = 0; i < boxes.length; i++) {
                boxes[i].checked = false;
            }
        }
        function updateToggleSelectAllLabel() {
            var btn = document.getElementById('toggle_select_all_btn');
            if (!btn) return;
            var boxes = document.getElementsByName('sales_select');
            var allChecked = true;
            if (!boxes || boxes.length === 0) { allChecked = false; }
            else {
                for (var i = 0; i < boxes.length; i++) { if (!boxes[i].checked) { allChecked = false; break; } }
            }
            btn.textContent = allChecked ? 'Αποεπιλογή όλων' : 'Επιλογή όλων';
        }
        function toggleSelectAllSales() {
            var boxes = document.getElementsByName('sales_select');
            if (!boxes || boxes.length === 0) return;
            var allChecked = true;
            for (var i = 0; i < boxes.length; i++) { if (!boxes[i].checked) { allChecked = false; break; } }
            var check = !allChecked;
            for (var j = 0; j < boxes.length; j++) { boxes[j].checked = check; }
            updateToggleSelectAllLabel();
        }
        function ensureDefaultAXSelection() {
            // If none is checked (first load), check all except the default excluded ones
            var boxes = document.getElementsByName('ax_filter');
            if (!boxes || boxes.length === 0) return;
            var anyChecked = false;
            for (var i = 0; i < boxes.length; i++) {
                if (boxes[i].checked) { anyChecked = true; break; }
            }
            if (!anyChecked) {
                var defEx = {{ default_excluded|tojson }};
                for (var i = 0; i < boxes.length; i++) {
                    var val = (boxes[i].value || '').trim();
                    if (defEx.indexOf(val) === -1) { boxes[i].checked = true; }
                }
            }
        }
        // ensure defaults after DOM is ready
        window.addEventListener('load', ensureDefaultAXSelection);

        function initStorageColumnVisibility() {
            var tbl = document.getElementById('materials_table');
            var chk = document.getElementById('toggle_storage_col');
            if (!tbl || !chk) return;
            var hidden = localStorage.getItem('hide_storage_col') === '1';
            if (hidden) {
                tbl.classList.add('hide-storage');
                chk.checked = false;
            } else {
                tbl.classList.remove('hide-storage');
                chk.checked = true;
            }
        }
        function onToggleStorageCol() {
            var tbl = document.getElementById('materials_table');
            var chk = document.getElementById('toggle_storage_col');
            if (!tbl || !chk) return;
            var hidden = !chk.checked;
            if (hidden) tbl.classList.add('hide-storage'); else tbl.classList.remove('hide-storage');
            localStorage.setItem('hide_storage_col', hidden ? '1' : '0');
        }
        window.addEventListener('load', initStorageColumnVisibility);

        function initExpectedByDateVisibility() {
            var tbl = document.getElementById('materials_table');
            var chk = document.getElementById('toggle_expected_bydate_col');
            if (!tbl || !chk) return;
            var hidden = localStorage.getItem('hide_expected_bydate_col') === '1';
            if (hidden) {
                tbl.classList.add('hide-expected-bydate');
                chk.checked = false;
            } else {
                tbl.classList.remove('hide-expected-bydate');
                chk.checked = true;
            }
        }
        function onToggleExpectedByDateCol() {
            var tbl = document.getElementById('materials_table');
            var chk = document.getElementById('toggle_expected_bydate_col');
            if (!tbl || !chk) return;
            var hidden = !chk.checked;
            if (hidden) tbl.classList.add('hide-expected-bydate'); else tbl.classList.remove('hide-expected-bydate');
            localStorage.setItem('hide_expected_bydate_col', hidden ? '1' : '0');
        }
        window.addEventListener('load', initExpectedByDateVisibility);

        function initUsedInVisibility() {
            var tbl = document.getElementById('materials_table');
            var chk = document.getElementById('toggle_usedin_col');
            if (!tbl || !chk) return;
            var hidden = localStorage.getItem('hide_usedin_col') === '1';
            if (hidden) {
                tbl.classList.add('hide-usedin');
                chk.checked = false;
            } else {
                tbl.classList.remove('hide-usedin');
                chk.checked = true;
            }
        }
        function onToggleUsedInCol() {
            var tbl = document.getElementById('materials_table');
            var chk = document.getElementById('toggle_usedin_col');
            if (!tbl || !chk) return;
            var hidden = !chk.checked;
            if (hidden) tbl.classList.add('hide-usedin'); else tbl.classList.remove('hide-usedin');
            localStorage.setItem('hide_usedin_col', hidden ? '1' : '0');
        }
        window.addEventListener('load', initUsedInVisibility);

        function initBring708Visibility() {
            var tbl = document.getElementById('materials_table');
            var chk = document.getElementById('toggle_bring708_col');
            if (!tbl || !chk) return;
            var hidden = localStorage.getItem('hide_bring708_col') === '1';
            if (hidden) {
                tbl.classList.add('hide-bring708');
                chk.checked = false;
            } else {
                tbl.classList.remove('hide-bring708');
                chk.checked = true;
            }
        }
        function onToggleBring708Col() {
            var tbl = document.getElementById('materials_table');
            var chk = document.getElementById('toggle_bring708_col');
            if (!tbl || !chk) return;
            var hidden = !chk.checked;
            if (hidden) tbl.classList.add('hide-bring708'); else tbl.classList.remove('hide-bring708');
            localStorage.setItem('hide_bring708_col', hidden ? '1' : '0');
        }
        window.addEventListener('load', initBring708Visibility);

        // ---------- Generic table sorting (text / number / date) ----------
        function parseNumberGR(s) {
            if (s === null || s === undefined) return 0;
            s = ("" + s).trim();
            if (s === '') return 0;
            // Remove thousand separators '.' and convert decimal comma to '.'
            s = s.replace(/\./g, '').replace(',', '.');
            var v = parseFloat(s);
            return isNaN(v) ? 0 : v;
        }

        // Trigger download: submit the existing form with a hidden flag
        function downloadExcel() {
            var f = document.querySelector('form');
            if (!f) return;
            var inp = document.createElement('input');
            inp.type = 'hidden';
            inp.name = 'download';
            inp.value = '1';
            f.appendChild(inp);
            f.submit();
            setTimeout(function(){ if (inp && inp.parentNode) inp.parentNode.removeChild(inp); }, 0);
        }

        // Trigger download of ALL rows (no 708 filter)
        function downloadExcelAll() {
            var f = document.querySelector('form');
            if (!f) return;
            var inp = document.createElement('input');
            inp.type = 'hidden';
            inp.name = 'download_all';
            inp.value = '1';
            f.appendChild(inp);
            f.submit();
            setTimeout(function(){ if (inp && inp.parentNode) inp.parentNode.removeChild(inp); }, 0);
        }

        // Open KR & Missing report in new page
        function openKrReport() {
            var f = document.querySelector('form');
            if (!f) return;
            var inp = document.createElement('input');
            inp.type = 'hidden';
            inp.name = 'kr_report';
            inp.value = '1';
            f.appendChild(inp);
            f.target = '_blank';
            f.submit();
            f.target = '';
            setTimeout(function(){ if (inp && inp.parentNode) inp.parentNode.removeChild(inp); }, 0);
        }

        // Hide rows where code starts with 4xx-xx-xx
        function applyHide4xx() {
            var tbl = document.getElementById('materials_table');
            var chk = document.getElementById('toggle_hide_4xx');
            if (!tbl || !chk) return;
            var hide = chk.checked;
            var rows = tbl.getElementsByTagName('tr');
            for (var i = 1; i < rows.length; i++) { // skip header
                var code = getCellText(rows[i], 0);
                var is4xx = /^\s*4\d{2}-\d{2}-\d{2}/.test(code || '');
                rows[i].style.display = (hide && is4xx) ? 'none' : '';
            }
        }
        function onToggleHide4xx() {
            var chk = document.getElementById('toggle_hide_4xx');
            if (!chk) return;
            localStorage.setItem('hide_4xx_rows', chk.checked ? '1' : '0');
            applyHide4xx();
        }
        function initHide4xx() {
            var chk = document.getElementById('toggle_hide_4xx');
            if (!chk) return;
            chk.checked = localStorage.getItem('hide_4xx_rows') === '1';
            applyHide4xx();
        }
        function getCellText(tr, idx) {
            var tds = tr.getElementsByTagName('td');
            if (idx < 0 || idx >= tds.length) return '';
            return (tds[idx].innerText || tds[idx].textContent || '').trim();
        }
        function sortTableBy(tableId, colIndex, type) {
            var table = document.getElementById(tableId);
            if (!table) return;
            var rows = Array.prototype.slice.call(table.querySelectorAll('tr'));
            if (rows.length === 0) return;
            var header = rows.shift(); // first row is header
            var currentCol = table.getAttribute('data-sort-col');
            var currentDir = table.getAttribute('data-sort-dir') || 'asc';
            var dir = (currentCol === String(colIndex) && currentDir === 'asc') ? 'desc' : 'asc';
            var modifier = dir === 'asc' ? 1 : -1;

            rows.sort(function(a, b){
                var A = getCellText(a, colIndex);
                var B = getCellText(b, colIndex);
                if (type === 'number') {
                    var na = parseNumberGR(A), nb = parseNumberGR(B);
                    if (na < nb) return -1 * modifier; if (na > nb) return 1 * modifier; return 0;
                }
                if (type === 'date') {
                    function parseDMY(s){
                        var m = (s||'').match(/^(\d{1,2})\/(\d{1,2})\/(\d{4})$/);
                        if (!m) return 0;
                        var d = parseInt(m[1],10), mo = parseInt(m[2],10)-1, y = parseInt(m[3],10);
                        return new Date(y, mo, d).getTime();
                    }
                    var da = parseDMY(A), db = parseDMY(B);
                    if (da < db) return -1 * modifier; if (da > db) return 1 * modifier; return 0;
                }
                // text
                A = (A||'').toLowerCase();
                B = (B||'').toLowerCase();
                var cmp = A.localeCompare(B, 'el', {numeric:true, sensitivity:'base'});
                return cmp * modifier;
            });

            // Reattach header then sorted rows
            var tbody = table.tBodies && table.tBodies[0] ? table.tBodies[0] : table;
            // If table has THEAD, keep using table element directly
            // Clear existing (except header)
            while (tbody.firstChild) tbody.removeChild(tbody.firstChild);
            tbody.appendChild(header);
            for (var i = 0; i < rows.length; i++) tbody.appendChild(rows[i]);

            table.setAttribute('data-sort-col', String(colIndex));
            table.setAttribute('data-sort-dir', dir);
            // Update header indicators
            var ths = header.getElementsByTagName('th');
            for (var j = 0; j < ths.length; j++) {
                if (ths[j].dataset && ths[j].dataset.sort && ths[j].dataset.sort !== 'none') {
                    ths[j].classList.add('sortable');
                    ths[j].removeAttribute('data-sorted');
                }
            }
            if (ths && ths[colIndex]) ths[colIndex].setAttribute('data-sorted', dir);
        }
        function initSortable(tableId) {
            var table = document.getElementById(tableId);
            if (!table) return;
            var header = table.querySelector('tr');
            if (!header) return;
            var ths = header.getElementsByTagName('th');
            for (var i = 0; i < ths.length; i++) {
                (function(idx){
                    var th = ths[idx];
                    var type = (th.dataset && th.dataset.sort) ? th.dataset.sort : 'text';
                    if (type === 'none') return;
                    th.classList.add('sortable');
                    th.addEventListener('click', function(){ sortTableBy(tableId, idx, type); });
                })(i);
            }
        }
        window.addEventListener('load', function(){
            initSortable('sales_table');
            initSortable('materials_table');
            initHide4xx();
        });
    </script>
</head>
<body>
<div class="container">
    <div class="page-header">
        <h2>Υπολογισμός Υλικών Παραγωγής</h2>
        <button type="button" onclick="openManualModal()" style="padding:7px 16px;">⚙️ Χειροκίνητα</button>
    </div>
    <form method="post">
        <input type="hidden" name="mode" id="mode_input" value="{{ mode }}">

        <!-- Filters modal overlay (inside form so checkboxes submit) -->
        <div id="filters_modal_overlay">
            <div id="filters_modal_box">
                <div class="modal-header">
                    <h3>Φίλτρα Αποθηκευτικών Χώρων</h3>
                    <button type="button" onclick="selectAllFilters()">Επιλογή όλων</button>
                    <button type="button" onclick="clearAllFilters()">Καθαρισμός</button>
                    <button type="button" id="filters_modal_close_btn" onclick="closeFiltersModal()" title="Κλείσιμο">&#10005;</button>
                </div>
                <div class="modal-body">
                    {% if storage_list %}
                    <div class="ax-filter-grid">
                        {% for ax in storage_list %}
                        <label>
                            <input type="checkbox" name="ax_filter" value="{{ ax }}" {% if ax in selected_ax %}checked{% endif %}>
                            {{ ax }}
                        </label>
                        {% endfor %}
                    </div>
                    {% else %}
                    <div>Δεν βρέθηκαν αποθηκευτικοί χώροι (Α.Χ.).</div>
                    {% endif %}
                </div>
                <div class="modal-footer">
                    <button type="button" onclick="closeFiltersModal()" style="padding:6px 18px;">Κλείσιμο</button>
                </div>
            </div>
        </div>
    <div class="tabs" style="display:none">
        <div class="tab" id="tab_manual" onclick="">Χειροκίνητα</div>
        <div class="tab active" id="tab_sales" onclick="">Από παραγγελίες</div>
    </div>

        <!-- Manual modal -->
        <div id="manual_modal_overlay">
            <div id="manual_modal_box">
                <div class="modal-header">
                    <h3>Χειροκίνητος Υπολογισμός</h3>
                    <button type="button" id="manual_modal_close_btn" onclick="closeManualModal()" title="Κλείσιμο">&#10005;</button>
                </div>
                <div class="modal-body">
                    <label>Κάθε γραμμή: Κωδικός Ποσότητα (π.χ. 321-00-16 10000)</label><br>
                    <textarea name="items" rows="10" style="width:100%; margin-top:8px;">{{ request.form['items'] if request.form.get('items') else '' }}</textarea>
                </div>
                <div class="modal-footer" style="display:flex; justify-content:space-between; align-items:center;">
                    <button type="button" onclick="closeManualModal()" style="padding:6px 18px;">Κλείσιμο</button>
                    <button type="submit" onclick="document.getElementById('mode_input').value='manual'" style="padding:6px 18px; font-weight:bold;">Υπολογισμός</button>
                </div>
            </div>
        </div>

    <div id="sales" style="display:block;">
            <div style="display:flex; gap:8px; align-items:center; margin-bottom:12px; flex-wrap:wrap;">
                <button type="submit" onclick="document.getElementById('mode_input').value='sales'" style="padding:8px 18px; font-weight:bold; background:#000; color:#fff; border:1px solid #000;">Υπολογισμός</button>
                <button type="button" id="toggle_orders_btn" onclick="toggleOrders()" style="padding:8px 14px;">Παραγγελίες πελατών</button>
                <button type="button" onclick="openFiltersModal()" style="padding:8px 14px;">Φίλτρα Α.Χ.</button>
            </div>

            <!-- Orders modal overlay (inside form for checkbox submission) -->
            <div id="orders_modal_overlay">
                <div id="orders_modal_box">
                    <div class="modal-header">
                        <h3>Παραγγελίες πελατών</h3>
                        <button type="button" id="toggle_select_all_btn" onclick="toggleSelectAllSales()">Επιλογή όλων</button>
                        <button type="button" onclick="clearSalesSelections(); updateToggleSelectAllLabel();">Καθαρισμός επιλογών</button>
                        <button type="button" id="orders_modal_close_btn" onclick="closeOrdersModal()" title="Κλείσιμο">&#10005;</button>
                    </div>
                    <div class="modal-body">
                        <input type="text" id="search" placeholder="Αναζήτηση..." style="width:100%;margin:0 0 10px 0;" onkeyup="filterSales()">
                        <table id="sales_table">
                            <tr>
                                <th data-sort="none">Επιλογή</th>
                                <th data-sort="text">Παραγγελία</th>
                                <th data-sort="text">Πελάτης</th>
                                <th data-sort="text">Είδος</th>
                                <th data-sort="number">Ανοικτή ποσότητα</th>
                                <th data-sort="date">Ημ/νία παράδοσης</th>
                                <th data-sort="date">Επιβεβαιωμένη ημερομηνία</th>
                            </tr>
                            {% for idx, row in sales_rows %}
                            <tr>
                                <td><input type="checkbox" name="sales_select" value="{{ idx }}" {% if request.form.getlist('sales_select') and (idx|string) in request.form.getlist('sales_select') %}checked{% endif %}></td>
                                <td>{{ row['Παραγγελία'] }}</td>
                                <td>{{ row['Πελάτης'] }}</td>
                                <td>{{ row['Είδος'] }}</td>
                                <td>{{ row['Ανοικτή ποσότητα'] }}</td>
                                <td>{{ row['Ημ/νία παράδοσης'] }}</td>
                                <td>{{ row['Επιβεβαιωμένη ημερομηνία'] }}</td>
                            </tr>
                            {% endfor %}
                        </table>
                    </div>
                    <div class="modal-footer">
                        <button type="button" onclick="closeOrdersModal()" style="padding:6px 18px;">Κλείσιμο</button>
                    </div>
                </div>
            </div>
        </div>
    </form>

    {% if details_agg %}
    <h3>Αναλυτικά είδη:</h3>
    <ul>
    {% for d in details_agg %}
        {% if d.not_found %}
            <li class="error">{{ d.product_code }} — Ποσότητα: {{ d.qty|gr }} — Δεν βρέθηκε συνταγή</li>
        {% else %}
            <li>{{ d.product_code }} ({{ d.description }}) [Έκδοση: /{{ d.max_rev }}] — Ποσότητα: {{ d.qty|gr }}</li>
        {% endif %}
    {% endfor %}
    </ul>
    {% endif %}

    {% if materials %}
    <h3>Αθροιστικά Υλικά:</h3>
    <div style="text-align:right; margin-bottom:6px; display:flex; gap:16px; justify-content:flex-end; align-items:center;">
        <label>
            <input type="checkbox" id="toggle_storage_col" onchange="onToggleStorageCol()" checked>
            Ανά αποθηκευτικό χώρο
        </label>
        <label>
            <input type="checkbox" id="toggle_expected_bydate_col" onchange="onToggleExpectedByDateCol()" checked>
            Αναμενόμενα ανά Ημ/νία 5 ή Παράδοση
        </label>
        <label>
            <input type="checkbox" id="toggle_usedin_col" onchange="onToggleUsedInCol()" checked>
            Αναλώνεται σε προϊόντα
        </label>
        <label>
            <input type="checkbox" id="toggle_bring708_col" onchange="onToggleBring708Col()" checked>
            Να φέρω από 708
        </label>
        <label>
            <input type="checkbox" id="toggle_hide_4xx" onchange="onToggleHide4xx()">
            Κρύψε 4xx-xx-xx
        </label>
        <button type="button" onclick="downloadExcel()" style="margin-left:12px; padding:6px 10px; border:1px solid #000; background:#eee; cursor:pointer;">📥 Κατέβασμα 708</button>
        <button type="button" onclick="downloadExcelAll()" style="margin-left:6px; padding:6px 10px; border:1px solid #555; background:#eee; cursor:pointer;">📥 Κατέβασμα όλων</button>
        <button type="button" onclick="openKrReport()" style="margin-left:12px; padding:6px 10px; border:1px solid #000; background:#ffeaa7; cursor:pointer;">ΚΡ & Ελλείψεις</button>
    </div>
    <table id="materials_table">
        <tr>
            <th data-sort="text">Κωδ. Υλικού</th>
            <th data-sort="text">Περιγραφή</th>
            <th class="num" data-sort="number">Ποσότητα</th>
            <th data-sort="text">Μονάδα</th>
            <th class="num" data-sort="number" title="χωρίς status 2">Στοκ</th>
            <th class="col-storage" data-sort="text">Ανά αποθηκευτικό χώρο</th>
            <th class="num" data-sort="number" title="ως προς Στοκ">Λείπει</th>
            <th class="num" data-sort="number" title="Αγορές">Αναμενόμενα</th>
            <th class="col-expected-bydate" data-sort="text">Αναμενόμενα ανά Ημ/νία 5 ή Παράδοση</th>
            <th class="num" data-sort="number">Προς παραγγελία</th>
            <th class="col-usedin" data-sort="text">Αναλώνεται σε προϊόντα</th>
            <th class="num col-bring708" data-sort="number">Να φέρω από 708</th>
        </tr>
        {% for m in materials %}
        <tr>
            <td>{{ m.code }}</td>
            <td>{{ m.desc }}</td>
            <td class="num">{{ m.qty|gr(2) }}</td>
            <td>{{ m.unit }}</td>
            <td class="num">{{ (m.stock|gr(2)) if m.stock is not none else '' }}</td>
            <td class="col-storage">{{ m.storage if m.storage else '' }}</td>
            <td class="num {% if m.missing_vs_stock and m.missing_vs_stock > 0 %}danger{% endif %}">{{ (m.missing_vs_stock|gr(2)) if m.missing_vs_stock is not none else '' }}</td>
            <td class="num">{{ (m.expected|gr(2)) if m.expected is not none else '' }}</td>
            <td class="col-expected-bydate">{{ m.expected_by_date if m.expected_by_date else '' }}</td>
            <td class="num {% if m.to_order and m.to_order > 0 %}danger{% endif %}">{{ (m.to_order|gr(2)) if m.to_order is not none else '' }}</td>
            <td class="col-usedin">{{ m.used_in if m.used_in else '' }}</td>
            <td class="num col-bring708 {% if m.bring_from_708_highlight %}danger{% endif %}">
                {{ (m.bring_from_708|gr(2)) if m.bring_from_708 is not none else '' }}
                {% if m.bring_from_708_detail %}<div style="font-size:0.9em; color:#555;">{{ m.bring_from_708_detail }}</div>{% endif %}
            </td>
        </tr>
        {% endfor %}
    </table>
    {% endif %}
</div>
</body>
</html>
"""


# Simple landing page with a single button linking to the app
HOME_TEMPLATE = """
<!doctype html>
<html>
<head>
    <meta charset="utf-8">
    <title>Αρχική</title>
    <style>
        body { background:#fff; color:#000; font-family:sans-serif; display:flex; align-items:center; justify-content:center; min-height:100vh; margin:0; padding:20px; box-sizing:border-box; }
        .card { border:1px solid #000; padding:24px 28px; text-align:center; max-width:480px; width:100%; }
        .btn { display:inline-block; padding:10px 16px; border:1px solid #000; background:#eee; color:#000; text-decoration:none; font-weight:600; }
        .btn:hover { background:#fff; }
        .upload-section { margin-top:20px; border-top:1px solid #ccc; padding-top:16px; }
        .upload-section h3 { margin-bottom:12px; font-size:1em; }
        .upload-row { display:flex; align-items:center; justify-content:space-between; margin-bottom:8px; gap:8px; }
        .upload-row label { font-size:0.85em; white-space:nowrap; min-width:110px; text-align:left; }
        .upload-row input[type=file] { font-size:0.8em; flex:1; min-width:0; }
        .upload-btn { margin-top:10px; padding:8px 20px; border:1px solid #000; background:#d4edda; cursor:pointer; font-weight:600; }
        .upload-btn:hover { background:#c3e6cb; }
        .msg-ok  { color:green; font-size:0.85em; margin-top:8px; }
        .msg-err { color:red;   font-size:0.85em; margin-top:8px; }
    </style>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <meta http-equiv="X-UA-Compatible" content="IE=edge" />
    <meta http-equiv="Content-Security-Policy" content="default-src 'self' 'unsafe-inline' data:;">
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    </head>
<body>
    <div class="card">
        <h2>Καλώς ήρθες</h2>
        <p>Προχώρα στον υπολογισμό των υλικών παραγωγής.</p>
        <p><a class="btn" href="{{ url_for('index') }}">Υπολογισμός Υλικών Παραγωγής</a></p>
    {# INACTIVE BUTTONS - uncomment to re-enable:
    <p><a class="btn" href="/usage">Ανάλωση υλικών & σχετικές παραγγελίες</a></p>
    <p><a class="btn" href="/procure-missing-date5">Αγορές χωρίς "Ημ/νία 5"</a></p>
    <p><a class="btn" href="/usage-by-storage">Ανάλωση ανά Α.Χ.</a></p>
    <p><a class="btn" href="/plan">Production Plan</a></p>
    <p><a class="btn" href="/semis">Ημιέτοιμα για Παραγωγή</a></p>
    #}

        <div class="upload-section">
            <h3>&#128196; Φόρτωση Αρχείων Δεδομένων</h3>
            {% if msg_ok %}<div class="msg-ok">&#10003; {{ msg_ok }}</div>{% endif %}
            {% if msg_err %}<div class="msg-err">&#9888; {{ msg_err }}</div>{% endif %}
            <form method="post" action="{{ url_for('upload_files') }}" enctype="multipart/form-data">
                <div class="upload-row">
                    <label>receipes.xlsx</label>
                    <input type="file" name="receipes" accept=".xlsx,.xls">
                </div>
                <div class="upload-row">
                    <label>sales.xlsx</label>
                    <input type="file" name="sales" accept=".xlsx,.xls">
                </div>
                <div class="upload-row">
                    <label>stock.xlsx</label>
                    <input type="file" name="stock" accept=".xlsx,.xls">
                </div>
                <div class="upload-row">
                    <label>procure.xlsx</label>
                    <input type="file" name="procure" accept=".xlsx,.xls">
                </div>
                <button type="submit" class="upload-btn">Ανέβασμα &amp; Ανανέωση</button>
            </form>
        </div>
    </div>
    </body>
    </html>
"""


@app.route('/')
def home():
    msg_ok  = request.args.get('msg_ok', '')
    msg_err = request.args.get('msg_err', '')
    return render_template_string(HOME_TEMPLATE, msg_ok=msg_ok, msg_err=msg_err)


@app.route('/upload', methods=['POST'])
def upload_files():
    global df, sales_df, stock_df, procure_df
    uploaded = []
    errors   = []

    file_map = {
        'receipes': 'receipes.xlsx',
        'sales':    'sales.xlsx',
        'stock':    'stock.xlsx',
        'procure':  'procure.xlsx',
    }

    for field, filename in file_map.items():
        f = request.files.get(field)
        if f and f.filename:
            try:
                f.save(filename)
                uploaded.append(filename)
            except Exception as e:
                errors.append(f"{filename}: {e}")

    # Reload dataframes after saving
    if 'receipes.xlsx' in uploaded:
        try:
            new_df = pd.read_excel('receipes.xlsx')
            if 'Ανενεργό' in new_df.columns:
                new_df = new_df[new_df['Ανενεργό'] == 0]
            df = new_df
        except Exception as e:
            errors.append(f'Σφάλμα φόρτωσης receipes.xlsx: {e}')

    if 'sales.xlsx' in uploaded:
        try:
            sales_df = pd.read_excel('sales.xlsx')
        except Exception as e:
            errors.append(f'Σφάλμα φόρτωσης sales.xlsx: {e}')

    if 'stock.xlsx' in uploaded:
        try:
            stock_df = pd.read_excel('stock.xlsx')
        except Exception as e:
            errors.append(f'Σφάλμα φόρτωσης stock.xlsx: {e}')

    if 'procure.xlsx' in uploaded:
        try:
            procure_df = pd.read_excel('procure.xlsx')
        except Exception as e:
            errors.append(f'Σφάλμα φόρτωσης procure.xlsx: {e}')

    msg_ok  = 'Ανέβηκαν: ' + ', '.join(uploaded) if uploaded else ''
    msg_err = 'Σφάλματα: ' + ' | '.join(errors)   if errors   else ''
    return redirect(url_for('home', msg_ok=msg_ok, msg_err=msg_err))


@app.route('/app', methods=['GET', 'POST'])
def index():
    materials = None
    details = None
    details_agg = None
    mode = request.form.get('mode', 'manual') if request.method == 'POST' else 'manual'

    # Prepare a stable subset for the sales table and selection (with confirmed date)
    sales_rows = []
    subset = pd.DataFrame()
    if not sales_df.empty:
        # Find columns robustly
        col_order = _find_col(sales_df, ["Παραγγελία", "Αρ. Παραγγελίας", "Παραγγελία Πώλησης", "Παραγγελία πελάτη"]) or "Παραγγελία"
        col_customer = _find_col(sales_df, ["Πελάτης", "Πελάτης/Όνομα", "Πελάτης-Όνομα"]) or "Πελάτης"
        col_item = _find_col(sales_df, ["Είδος", "Κωδικός Είδους", "Κωδ. Είδους"]) or "Είδος"
        col_open = _find_col(sales_df, ["Ανοικτή ποσότητα", "Υπόλοιπο", "Υπόλοιπο Ποσότητας", "Υπολ. Ποσότητα"]) or "Ανοικτή ποσότητα"
        col_delivery = _find_col(sales_df, ["Ημ/νία παράδοσης", "Ημερομηνία παράδοσης", "Παράδοση"]) or "Ημ/νία παράδοσης"
        col_confirmed = _find_col(sales_df, ["Επιβεβαιωμένη ημερομηνία", "Επιβεβαιωμένη", "Επιβεβαιωμ"])  # may be None

        cols_to_take = [c for c in [col_order, col_customer, col_item, col_open, col_delivery] if c in sales_df.columns]
        if col_confirmed and col_confirmed in sales_df.columns:
            cols_to_take.append(col_confirmed)
        subset = sales_df[cols_to_take].copy() if cols_to_take else sales_df.copy()

        # Normalize column names to our display keys
        rename_map = {}
        if col_order in subset.columns: rename_map[col_order] = "Παραγγελία"
        if col_customer in subset.columns: rename_map[col_customer] = "Πελάτης"
        if col_item in subset.columns: rename_map[col_item] = "Είδος"
        if col_open in subset.columns: rename_map[col_open] = "Ανοικτή ποσότητα"
        if col_delivery in subset.columns: rename_map[col_delivery] = "Ημ/νία παράδοσης"
        if col_confirmed and col_confirmed in subset.columns:
            rename_map[col_confirmed] = "Επιβεβαιωμένη ημερομηνία"
        subset = subset.rename(columns=rename_map)

        # Ensure both date columns are formatted as date-only strings (dd/mm/yyyy)
        for date_col in ["Ημ/νία παράδοσης", "Επιβεβαιωμένη ημερομηνία"]:
            if date_col in subset.columns:
                s = pd.to_datetime(subset[date_col], errors='coerce')
                subset[date_col] = s.dt.strftime('%d/%m/%Y').fillna('')
        # Quantity numeric
        if "Ανοικτή ποσότητα" in subset.columns:
            subset["Ανοικτή ποσότητα"] = pd.to_numeric(subset["Ανοικτή ποσότητα"], errors='coerce').fillna(0)

        # If confirmed column missing, add empty one for template stability
        if "Επιβεβαιωμένη ημερομηνία" not in subset.columns:
            subset["Επιβεβαιωμένη ημερομηνία"] = ""

        subset = subset.head(500)
        sales_rows = list(enumerate(subset.to_dict(orient='records')))

    # Prepare A.X. filter values
    storage_list = get_storage_list(stock_df)
    default_excluded = ["06", "Σ06", "23", "26"]
    if request.method == 'POST':
        selected_ax = request.form.getlist('ax_filter')
    else:
        # Default: select all except the specified storages
        selected_ax = [ax for ax in storage_list if ax not in default_excluded]

    if request.method == 'POST':
        items = []
        if mode == 'manual':
            lines = request.form.get('items', '').splitlines()
            for line in lines:
                parts = line.strip().split()
                if len(parts) >= 2:
                    code = parts[0]
                    try:
                        qty = float(parts[1])
                    except Exception:
                        qty = 0
                    items.append({'product_code': code, 'qty': qty})
        elif mode == 'sales' and not subset.empty:
            selected = request.form.getlist('sales_select')
            for idx in selected:
                try:
                    pos = int(idx)
                except Exception:
                    continue
                # Map back to the same subset row order
                row = subset.iloc[pos]
                code = str(row.get('Είδος', '')).split()[0]
                qty = row.get('Ανοικτή ποσότητα', 0)
                try:
                    qty = float(qty)
                except Exception:
                    qty = 0
                items.append({'product_code': code, 'qty': qty})

        if items:
            materials, details = get_materials_multiple(items)
            # Aggregate requested quantities per top-level product
            qty_map = {}
            for it in items:
                base = _base_code(str(it.get('product_code','')))
                try:
                    q = float(it.get('qty', 0) or 0)
                except Exception:
                    q = 0
                if base:
                    qty_map[base] = qty_map.get(base, 0.0) + q
            # Deduplicate details and attach qty
            details_agg = []
            seen = set()
            for d in details or []:
                base = _base_code(d.get('product_code',''))
                if base in seen:
                    continue
                seen.add(base)
                details_agg.append({
                    'product_code': base,
                    'description': d.get('description',''),
                    'max_rev': d.get('max_rev',''),
                    'not_found': d.get('not_found', False),
                    'qty': qty_map.get(base, 0.0)
                })
            # Build mapping: material -> set of top-level products where it is used
            used_in_map = {}
            prod_lookup = {str(d.get('product_code','')): str(d.get('description','')) for d in (details or []) if not d.get('not_found')}
            cache_used = {}
            for it in items:
                pcode = str(it.get('product_code',''))
                base = _base_code(pcode)
                per_unit = _per_unit_materials(base, cache_used, set())
                disp = base
                desc = prod_lookup.get(base, '')
                if desc:
                    disp = f"{base} ({desc})"
                for k in per_unit.keys():
                    s = used_in_map.setdefault(k, set())
                    s.add(disp)
            # Enrich with stock and expected procure info if available
            stock_index = build_stock_index(stock_df)
            procure_index = build_procure_index(procure_df)
            procure_by_date = build_procure_by_date(procure_df)
            stock_lots = build_stock_lots(stock_df)
            for m in materials:
                code = str(m.get('code', '')).strip()
                # Stock enrichment
                info = stock_index.get(code) if stock_index else None
                if info:
                    by = (info.get('by_storage', {}) or {})
                    # Apply A.X. filters: include only checked storages
                    if selected_ax:
                        filtered_by = {k: v for k, v in by.items() if k in selected_ax}
                    else:
                        filtered_by = {}
                    m['stock'] = round(sum(filtered_by.values()), 3)
                    parts = [f"{k}: {format_gr(v, 2)}" for k, v in filtered_by.items()]
                    m['storage'] = ", ".join(parts)
                else:
                    m['stock'] = None
                    m['storage'] = ''
                # Procure enrichment (expected incoming)
                if procure_index and code in procure_index:
                    m['expected'] = round(procure_index.get(code, 0.0), 3)
                else:
                    m['expected'] = None
                # Procure by date (Ημ/νία 5)
                if procure_by_date and code in procure_by_date:
                    bydate = procure_by_date.get(code, {})
                    def _to_dmy(dstr):
                        try:
                            dt = pd.to_datetime(dstr, errors='coerce')
                            return dt.strftime('%d/%m/%Y') if pd.notna(dt) else str(dstr)
                        except Exception:
                            return str(dstr)
                    parts = [f"{_to_dmy(d)}: {format_gr(qty, 2)}" for d, qty in sorted(bydate.items())]
                    m['expected_by_date'] = ", ".join(parts)
                else:
                    m['expected_by_date'] = ''
                # Used in (top-level products)
                key = (m.get('code'), m.get('desc'), m.get('unit'))
                if used_in_map.get(key):
                    m['used_in'] = ", ".join(sorted(used_in_map[key]))
                else:
                    m['used_in'] = ''
                # Shortage / to order
                req = float(m.get('qty', 0) or 0)
                stk = float(m.get('stock', 0) or 0)
                exp = float(m.get('expected', 0) or 0)
                missing_vs_stock = req - stk
                m['missing_vs_stock'] = round(missing_vs_stock, 3) if missing_vs_stock > 0 else 0.0
                to_order = req - stk - exp
                m['to_order'] = round(to_order, 3) if to_order > 0 else 0.0

                # New: Global FEFO (excluding '07') and compute how much to bring from 708 with details/highlight
                bring708 = 0.0
                bring708_detail = []
                bring708_highlight = False
                lots = stock_lots.get(code, []) if stock_lots else []
                if lots:
                    def expiry_key(x):
                        e = x.get('expiry')
                        return (e is None, e if isinstance(e, pd.Timestamp) else pd.Timestamp.max)
                    fefo_lots = [x.copy() for x in lots if str(x.get('storage')) != '07']
                    fefo_lots.sort(key=expiry_key)
                    req_left = req
                    picked_from_others_total = 0.0
                    for lot in fefo_lots:
                        if req_left <= 0:
                            break
                        avail = float(lot.get('qty', 0) or 0)
                        if avail <= 0:
                            continue
                        take = min(req_left, avail)
                        req_left -= take
                        if str(lot.get('storage')) == '708':
                            bring708 += take
                            let = lot.get('expiry')
                            try:
                                dstr = let.strftime('%d/%m/%Y') if isinstance(let, pd.Timestamp) and pd.notna(let) else ''
                            except Exception:
                                dstr = ''
                            b = str(lot.get('batch') or '')
                            part = f"{('['+b+'] ') if b else ''}{dstr}: {format_gr(take, 2)}"
                            bring708_detail.append(part)
                        else:
                            picked_from_others_total += take
                    # Highlight if FEFO decided to use 708 while other storages also had stock
                    others_total = sum(float(x.get('qty', 0) or 0) for x in fefo_lots if str(x.get('storage')) != '708')
                    if bring708 > 0 and others_total > picked_from_others_total:
                        bring708_highlight = True
                m['bring_from_708'] = round(bring708, 3) if bring708 > 0 else 0.0
                m['bring_from_708_detail'] = ", ".join(bring708_detail) if bring708_detail else ''
                m['bring_from_708_highlight'] = bring708_highlight

            # If user requested KR & Missing report, build and return it
            if request.form.get('kr_report') == '1' and materials:
                return render_kr_missing_report(materials)

            # If user requested an Excel download, build and return it
            if (request.form.get('download') == '1' or request.form.get('download_all') == '1') and materials:
                only_708 = request.form.get('download') == '1'
                export_rows = []
                for m in materials:
                    export_rows.append({
                        'Κωδ. Υλικού': m.get('code'),
                        'Περιγραφή': m.get('desc'),
                        'Ποσότητα': m.get('qty'),
                        'Μονάδα': m.get('unit'),
                        'Στοκ (χωρίς status 2)': m.get('stock'),
                        'Ανά αποθηκευτικό χώρο': m.get('storage'),
                        'Λείπει (ως προς Στοκ)': m.get('missing_vs_stock'),
                        'Αναμενόμενα (Αγορές)': m.get('expected'),
                        'Αναμενόμενα ανά Ημ/νία 5 ή Παράδοση': m.get('expected_by_date'),
                        'Προς παραγγελία': m.get('to_order'),
                        'Αναλώνεται σε προϊόντα': m.get('used_in'),
                        'Να φέρω από 708': m.get('bring_from_708'),
                        'Λεπτομέρειες 708': m.get('bring_from_708_detail'),
                    })
                out_df = pd.DataFrame(export_rows)
                # Order columns similar to the UI
                cols = [
                    'Κωδ. Υλικού','Περιγραφή','Ποσότητα','Μονάδα','Στοκ (χωρίς status 2)','Ανά αποθηκευτικό χώρο',
                    'Λείπει (ως προς Στοκ)','Αναμενόμενα (Αγορές)','Αναμενόμενα ανά Ημ/νία 5 ή Παράδοση','Προς παραγγελία',
                    'Αναλώνεται σε προϊόντα','Να φέρω από 708','Λεπτομέρειες 708'
                ]
                out_df = out_df[[c for c in cols if c in out_df.columns]]
                # Filter to only rows with 708 quantity when requested
                if only_708 and 'Να φέρω από 708' in out_df.columns:
                    out_df = out_df[out_df['Να φέρω από 708'].fillna(0) > 0]
                fname_suffix = '708' if only_708 else 'all'

                buf = io.BytesIO()
                # Try xlsxwriter for better compatibility, fallback to openpyxl
                try:
                    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
                        out_df.to_excel(writer, index=False, sheet_name='Materials')
                except Exception:
                    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                        out_df.to_excel(writer, index=False, sheet_name='Materials')
                buf.seek(0)
                return send_file(
                    buf,
                    as_attachment=True,
                    download_name=f"materials_{fname_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

    return render_template_string(TEMPLATE, materials=materials, details=details, details_agg=details_agg, request=request, sales_rows=sales_rows, mode=mode, storage_list=storage_list, selected_ax=selected_ax, default_excluded=default_excluded)

def render_kr_missing_report(materials):
    """Render KR & Missing materials report"""
    
    # Build stock lots index for KR details
    stock_lots = {}
    try:
        if not stock_df.empty:
            stock_lots = build_stock_lots(stock_df)
    except Exception:
        pass
    
    def extract_kr_stock_with_batches(code, storage_text):
        """Extract KR stock quantity and batch info from storage text and lots"""
        if not storage_text or 'ΚΡ' not in storage_text:
            return ''
        
        # Get lots for this code
        lots = stock_lots.get(code, [])
        kr_lots = [lot for lot in lots if str(lot.get('storage', '')).strip() == 'ΚΡ']
        
        if kr_lots:
            # Format batch info: "Batch1: 100.00, Batch2: 50.00"
            batch_details = []
            total_qty = 0
            for lot in kr_lots:
                batch = lot.get('batch', '') or 'N/A'
                qty = float(lot.get('qty', 0) or 0)
                total_qty += qty
                if batch != 'N/A':
                    batch_details.append(f"{batch}: {format_gr(qty, 2)}")
                else:
                    batch_details.append(f"{format_gr(qty, 2)}")
            
            if batch_details:
                return "\n".join(batch_details)
            else:
                return format_gr(total_qty, 2) if total_qty > 0 else ''
        
        # Fallback to parsing storage text
        import re
        kr_match = re.search(r'ΚΡ[:\s]*([0-9.,]+)', storage_text)
        if kr_match and kr_match.group(1):
            return kr_match.group(1).replace(',', '.')
        
        return 'ΚΡ'
    
    # Filter materials with KR or missing > 0
    filtered_materials = []
    for m in materials:
        storage = m.get('storage', '') or ''
        missing = float(m.get('missing_vs_stock', 0) or 0)
        
        has_kr = 'ΚΡ' in storage
        has_missing = missing > 0
        
        if has_kr or has_missing:
            kr_stock = extract_kr_stock_with_batches(m.get('code', ''), storage)
            filtered_materials.append({
                'code': m.get('code', ''),
                'desc': m.get('desc', ''),
                'qty': m.get('qty', ''),
                'unit': m.get('unit', ''),
                'stock': m.get('stock', ''),
                'kr_stock': kr_stock,
                'missing': m.get('missing_vs_stock', ''),
                'has_kr': has_kr,
                'has_missing': has_missing
            })
    
    return render_template_string(KR_MISSING_TEMPLATE, materials=filtered_materials)

@app.route('/kr-missing-excel', methods=['POST'])
def kr_missing_excel():
    """Download KR & Missing report as Excel"""
    import json
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file
    
    try:
        # Get materials data from form
        materials_json = request.form.get('materials_data', '[]')
        materials = json.loads(materials_json)
        
        if not materials:
            return "Δεν υπάρχουν δεδομένα για εξαγωγή", 400
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ΚΡ & Ελλείψεις"
        
        # Headers
        headers = [
            'Κωδικός Υλικού', 'Περιγραφή', 'Απαίτηση', 'Μονάδα', 
            'Στοκ', 'Στοκ σε ΚΡ', 'Αναμενόμενο'
        ]
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center')
        right_align = Alignment(horizontal='right')
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Write data
        for row_idx, material in enumerate(materials, 2):
            # Clean numeric values
            def clean_numeric(val):
                if val == '-' or not val:
                    return ''
                # Remove Greek number formatting
                cleaned = str(val).replace('.', '').replace(',', '.')
                try:
                    return float(cleaned)
                except:
                    return val
            
            values = [
                material.get('code', ''),
                material.get('desc', ''),
                clean_numeric(material.get('qty', '')),
                material.get('unit', ''),
                clean_numeric(material.get('stock', '')),
                material.get('kr_stock', ''),  # Keep as text for batch info
                clean_numeric(material.get('missing', ''))
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                # Right align numeric columns
                if col in [3, 5, 7]:  # Qty, Stock, Missing columns
                    cell.alignment = right_align
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate filename with current date
        from datetime import datetime
        current_date = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"KR_Ellipseis_{current_date}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return f"Σφάλμα κατά τη δημιουργία του Excel: {str(e)}", 500

KR_MISSING_TEMPLATE = r"""
<!doctype html>
<html>
<head>
    <meta charset="utf-8">
    <title>ΚΡ & Ελλείψεις</title>
    <style>
        body { background:#fff; color:#000; font-family:sans-serif; margin:20px; }
        h1 { font-size: 20px; margin-bottom: 15px; border-bottom: 2px solid #333; }
        table { border-collapse: collapse; width: 100%; margin-top: 15px; }
        th, td { border: 1px solid #333; padding: 8px; text-align: left; }
        th { background-color: #f0f0f0; font-weight: bold; }
        .num { text-align: right; font-variant-numeric: tabular-nums; }
        .kr-row { background-color: #e8f5e8; }
        .missing-row { background-color: #ffe8e8; }
        .both-row { background-color: #fff2e8; }
        .summary { margin-top: 20px; padding: 10px; background-color: #f9f9f9; border: 1px solid #ddd; }
        .btn { display: inline-block; padding: 8px 12px; margin: 10px 5px 0 0; border: 1px solid #333; background: #eee; color: #333; text-decoration: none; }
        .btn:hover { background: #ddd; }
        .kr-batches { font-size: 11px; color: #555; white-space: pre-wrap; }
    </style>
    <script>
        function downloadKrExcel() {
            var form = document.createElement('form');
            form.method = 'POST';
            form.action = '/kr-missing-excel';
            form.target = '_blank';
            
            var materialsInput = document.createElement('input');
            materialsInput.type = 'hidden';
            materialsInput.name = 'materials_data';
            
            // Collect table data
            var tableData = [];
            var rows = document.querySelectorAll('tbody tr');
            rows.forEach(function(row) {
                var cells = row.querySelectorAll('td');
                if (cells.length >= 7) {
                    tableData.push({
                        code: cells[0].textContent.trim(),
                        desc: cells[1].textContent.trim(),
                        qty: cells[2].textContent.trim(),
                        unit: cells[3].textContent.trim(),
                        stock: cells[4].textContent.trim(),
                        kr_stock: cells[5].textContent.trim(),
                        missing: cells[6].textContent.trim()
                    });
                }
            });
            
            materialsInput.value = JSON.stringify(tableData);
            form.appendChild(materialsInput);
            
            document.body.appendChild(form);
            form.submit();
            document.body.removeChild(form);
        }
    </script>
</head>
<body>
    <h1>Αναφορά ΚΡ & Ελλείψεων</h1>
    <p><strong>Ημερομηνία:</strong> <script>document.write(new Date().toLocaleDateString('el-GR') + ' ' + new Date().toLocaleTimeString('el-GR'));</script></p>
    
    <a href="{{ url_for('index') }}" class="btn">← Επιστροφή</a>
    <button onclick="window.print()" class="btn">🖨 Εκτύπωση</button>
    <button onclick="downloadKrExcel()" class="btn">📊 Κατέβασμα σε Excel</button>
    
    {% if materials %}
    <table>
        <thead>
            <tr>
                <th>Κωδικός Υλικού</th>
                <th>Περιγραφή</th>
                <th class="num">Απαίτηση</th>
                <th>Μονάδα</th>
                <th class="num">Στοκ</th>
                <th class="num">Στοκ σε ΚΡ</th>
                <th class="num">Αναμενόμενο</th>
            </tr>
        </thead>
        <tbody>
            {% for m in materials %}
            {% set row_class = 'both-row' if m.has_kr and m.has_missing else ('kr-row' if m.has_kr else 'missing-row') %}
            <tr class="{{ row_class }}">
                <td>{{ m.code }}</td>
                <td>{{ m.desc }}</td>
                <td class="num">{{ m.qty|gr(2) if m.qty else '-' }}</td>
                <td>{{ m.unit }}</td>
                <td class="num">{{ m.stock|gr(2) if m.stock else '-' }}</td>
                <td class="num kr-batches">{{ m.kr_stock if m.kr_stock else '-' }}</td>
                <td class="num">{{ m.missing|gr(2) if m.has_missing else '-' }}</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
    
    <div class="summary">
        <strong>Σύνοψη:</strong><br>
        Συνολικά είδη: {{ materials|length }}<br>
        Με στοκ σε ΚΡ: {{ materials|selectattr('has_kr')|list|length }}<br>
        Με ελλείψεις: {{ materials|selectattr('has_missing')|list|length }}<br>
        Και τα δύο: {{ materials|selectattr('has_kr')|selectattr('has_missing')|list|length }}
    </div>
    {% else %}
    <p>Δεν βρέθηκαν υλικά με ΚΡ ή ελλείψεις.</p>
    {% endif %}
</body>
</html>
"""

USAGE_TEMPLATE = r"""
<!doctype html>
<html>
<head>
    <title>Ανάλωση υλικών & σχετικές παραγγελίες</title>
    <style>
        body { background:#fff; color:#000; font-family:sans-serif; margin:0; }
        .container { width:100%; padding:16px 20px; box-sizing:border-box; }
        textarea { width:100%; height:110px; }
        table { width:100%; border-collapse: collapse; margin-top: 12px; }
        th, td { border:1px solid #000; padding:4px; text-align:left; }
        .num { text-align:right; font-variant-numeric: tabular-nums; }
        .btn { display:inline-block; padding:8px 12px; border:1px solid #000; background:#eee; color:#000; text-decoration:none; cursor:pointer; }
    </style>
    <script>
        function downloadUsageExcel() {
            var f = document.querySelector('form');
            if (!f) return;
            var inp = document.createElement('input');
            inp.type = 'hidden';
            inp.name = 'download';
            inp.value = '1';
            f.appendChild(inp);
            f.submit();
            setTimeout(function(){ if (inp && inp.parentNode) inp.parentNode.removeChild(inp); }, 0);
        }
        // Sorting for tables on this page
        function parseNumberGR(s){ if(s==null) return 0; s=(''+s).trim(); if(s==='') return 0; s=s.replace(/\./g,'').replace(',', '.'); var v=parseFloat(s); return isNaN(v)?0:v; }
        function parseDMY(s){ var m=(''+(s||'')).match(/^(\d{1,2})\/(\d{1,2})\/(\d{4})$/); if(!m) return 0; return new Date(parseInt(m[3],10), parseInt(m[2],10)-1, parseInt(m[1],10)).getTime(); }
        function sortTableGeneric(table, colIndex, type){
            var rows=Array.prototype.slice.call(table.querySelectorAll('tr'));
            if(rows.length<2) return; var header=rows.shift();
            var currentCol=table.getAttribute('data-sort-col'); var currentDir=table.getAttribute('data-sort-dir')||'asc';
            var dir=(currentCol===String(colIndex)&&currentDir==='asc')?'desc':'asc'; var mod=dir==='asc'?1:-1;
            rows.sort(function(a,b){
                var A=(a.getElementsByTagName('td')[colIndex]||{}).innerText||'';
                var B=(b.getElementsByTagName('td')[colIndex]||{}).innerText||'';
                if(type==='number'){ var na=parseNumberGR(A), nb=parseNumberGR(B); if(na<nb) return -1*mod; if(na>nb) return 1*mod; return 0; }
                if(type==='date'){ var da=parseDMY(A), db=parseDMY(B); if(da<db) return -1*mod; if(da>db) return 1*mod; return 0; }
                A=A.toLowerCase(); B=B.toLowerCase(); var cmp=A.localeCompare(B,'el',{numeric:true,sensitivity:'base'}); return cmp*mod;
            });
            while (table.firstChild) table.removeChild(table.firstChild);
            table.appendChild(header); for (var i=0;i<rows.length;i++) table.appendChild(rows[i]);
            table.setAttribute('data-sort-col', String(colIndex)); table.setAttribute('data-sort-dir', dir);
            var ths=header.getElementsByTagName('th'); for (var j=0;j<ths.length;j++){ ths[j].removeAttribute('data-sorted'); ths[j].style.cursor='pointer'; }
            if (ths[colIndex]) ths[colIndex].setAttribute('data-sorted', dir);
        }
        function initSortableTables(){
            var tables = document.querySelectorAll('table.usage-sortable');
            for (var t=0; t<tables.length; t++){
                (function(tbl){ var header=tbl.querySelector('tr'); if(!header) return; var ths=header.getElementsByTagName('th'); for (var i=0;i<ths.length;i++){ (function(idx){ var th=ths[idx]; var type=th.getAttribute('data-sort')||'text'; if(type==='none') return; th.style.cursor='pointer'; th.onclick=function(){ sortTableGeneric(tbl, idx, type); }; })(i); } })(tables[t]);
            }
        }
        window.addEventListener('load', initSortableTables);
    </script>
</head>
<body>
    <div class="container">
        <h2>Ανάλωση υλικών & σχετικές παραγγελίες</h2>
        <form method="post">
            <p>Δώσε κωδικούς αναλούμενων (ένας ανά γραμμή):</p>
            <textarea name="materials" placeholder="π.χ. 224-01-87\n224-04-45">{{ request.form['materials'] if request.form.get('materials') else '' }}</textarea>
            <div>
                <button class="btn" type="submit">Ανάλυση</button>
                <a class="btn" href="/">Αρχική</a>
                {% if results %}
                <button class="btn" type="button" onclick="downloadUsageExcel()">Κατέβασμα σε Excel</button>
                {% endif %}
            </div>
        </form>

        {% if results %}
        <h3>Αποτελέσματα</h3>
        {% for r in results %}
            <h4>{{ r.leaf }} {% if r.desc %}— {{ r.desc }}{% endif %}</h4>
            {% if r.products %}
            <p>Αναλώνεται στα παρακάτω προϊόντα:</p>
            <ul>
                {% for p in r.products %}
                    <li>
                        {{ p.code }}{% if p.desc %} — {{ p.desc }}{% endif %}
                        {% if p.per_unit is not none and p.unit %}
                            — ανά τεμάχιο: {{ p.per_unit|gr(2) }} {{ p.unit }}
                        {% endif %}
                    </li>
                {% endfor %}
            </ul>
            {% else %}
            <p>Δεν βρέθηκαν προϊόντα που το αναλώνουν.</p>
            {% endif %}

            {% if r.orders and r.orders|length > 0 %}
            <table class="usage-sortable">
                <tr>
                    <th data-sort="text">Παραγγελία</th>
                    <th data-sort="text">Πελάτης</th>
                    <th data-sort="text">Είδος</th>
                    <th class="num" data-sort="number">Ανοικτή ποσότητα</th>
                    <th data-sort="date">Ημ/νία παράδοσης</th>
                    <th data-sort="date">Επιβεβαιωμένη ημερομηνία</th>
                </tr>
                {% for o in r.orders %}
                <tr>
                    <td>{{ o.order }}</td>
                    <td>{{ o.customer }}</td>
                    <td>{{ o.item }}</td>
                    <td class="num">{{ o.open|gr(2) }}</td>
                    <td>{{ o.delivery }}</td>
                    <td>{{ o.confirmed }}</td>
                </tr>
                {% endfor %}
            </table>
            {% else %}
            <p>Δεν βρέθηκαν σχετικές ανοιχτές παραγγελίες.</p>
            {% endif %}
            <hr>
        {% endfor %}
        {% endif %}
    </div>
</body>
</html>
"""

SEMI_TEMPLATE = r"""
<!doctype html>
<html>
<head>
    <meta charset="utf-8">
    <title>Ημιέτοιμα για Παραγωγή</title>
    <style>
        body { background:#fff; color:#000; font-family:sans-serif; margin:0; }
        .container { width:100%; padding:16px 20px; box-sizing:border-box; }
        textarea { width:100%; height:130px; }
        table { width:100%; border-collapse: collapse; margin-top: 12px; }
        th, td { border:1px solid #000; padding:4px; text-align:left; }
        .num { text-align:right; font-variant-numeric: tabular-nums; }
        .btn { display:inline-block; padding:8px 12px; border:1px solid #000; background:#eee; color:#000; text-decoration:none; cursor:pointer; }
        .btn:hover { background:#ddd; }
        th[data-sort] { cursor:pointer; }
    </style>
    <script>
        function parseNumberGR(s){ if(s==null) return 0; s=(''+s).trim(); if(s==='') return 0; s=s.replace(/\./g,'').replace(',', '.'); var v=parseFloat(s); return isNaN(v)?0:v; }
        function sortTableGeneric(table, colIndex, type){
            var rows=Array.prototype.slice.call(table.querySelectorAll('tr'));
            if(rows.length<2) return; var header=rows.shift();
            var currentCol=table.getAttribute('data-sort-col'); var currentDir=table.getAttribute('data-sort-dir')||'asc';
            var dir=(currentCol===String(colIndex)&&currentDir==='asc')?'desc':'asc'; var mod=dir==='asc'?1:-1;
            rows.sort(function(a,b){
                var A=(a.getElementsByTagName('td')[colIndex]||{}).innerText||'';
                var B=(b.getElementsByTagName('td')[colIndex]||{}).innerText||'';
                if(type==='number'){ var na=parseNumberGR(A), nb=parseNumberGR(B); if(na<nb) return -1*mod; if(na>nb) return 1*mod; return 0; }
                A=A.toLowerCase(); B=B.toLowerCase(); var cmp=A.localeCompare(B,'el',{numeric:true,sensitivity:'base'}); return cmp*mod;
            });
            while (table.firstChild) table.removeChild(table.firstChild);
            table.appendChild(header); for (var i=0;i<rows.length;i++) table.appendChild(rows[i]);
            table.setAttribute('data-sort-col', String(colIndex)); table.setAttribute('data-sort-dir', dir);
            var ths=header.getElementsByTagName('th'); for (var j=0;j<ths.length;j++){ ths[j].removeAttribute('data-sorted'); }
            if (ths[colIndex]) ths[colIndex].setAttribute('data-sorted', dir);
        }
        function initSortable(){
            var tbl=document.getElementById('semis_table'); if(!tbl) return;
            var header=tbl.querySelector('tr'); if(!header) return;
            var ths=header.getElementsByTagName('th');
            for(var i=0;i<ths.length;i++){
                (function(idx){ var th=ths[idx]; var t=th.getAttribute('data-sort')||'text'; if(t==='none') return; th.onclick=function(){ sortTableGeneric(tbl, idx, t); }; })(i);
            }
        }
        window.addEventListener('load', initSortable);
        function downloadSemisExcel(){
            var f=document.getElementById('semis_form'); if(!f) return;
            var inp=document.createElement('input'); inp.type='hidden'; inp.name='download'; inp.value='1'; f.appendChild(inp);
            f.submit(); setTimeout(function(){ if(inp && inp.parentNode) inp.parentNode.removeChild(inp); }, 0);
        }
    </script>
</head>
<body>
    <div class="container">
        <h2>Ημιέτοιμα για Παραγωγή</h2>
        <form method="post" id="semis_form">
            <p>Δώσε κωδικούς προϊόντων με ποσότητα (ένας ανά γραμμή, π.χ. <code>121-00-01 500</code>):</p>
            <textarea name="items" placeholder="π.χ. 121-00-01 500\n122-05-10 200">{{ request.form['items'] if request.form.get('items') else '' }}</textarea>
            <div style="margin-top:8px;">
                <button class="btn" type="submit">Υπολογισμός</button>
                <a class="btn" href="/">Αρχική</a>
                {% if semis %}
                <button class="btn" type="button" onclick="downloadSemisExcel()">Κατέβασμα σε Excel</button>
                {% endif %}
            </div>
        </form>

        {% if details %}
        <h3>Επιλεγμένα προϊόντα:</h3>
        <ul>
        {% for d in details %}
            {% if d.not_found %}
                <li class="error">{{ d.product_code }} — Δεν βρέθηκε συνταγή</li>
            {% else %}
                <li>{{ d.product_code }} ({{ d.description }}) [Έκδοση /{{ d.max_rev }}]</li>
            {% endif %}
        {% endfor %}
        </ul>
        {% endif %}

        {% if combined %}
        <h3>Αθροιστικά: Προϊόντα & Ημιέτοιμα</h3>
        <table id="semis_table">
            <tr>
                <th data-sort="text">Τύπος</th>
                <th data-sort="text">Κωδικός</th>
                <th data-sort="text">Περιγραφή</th>
                <th class="num" data-sort="number">Ποσότητα</th>
                <th data-sort="text">Μονάδα</th>
            </tr>
            {% for r in combined %}
            <tr>
                <td>{{ r.type }}</td>
                <td>{{ r.code }}</td>
                <td>{{ r.desc }}</td>
                <td class="num">{{ r.qty|gr(3) }}</td>
                <td>{{ r.unit }}</td>
            </tr>
            {% endfor %}
        </table>
        {% endif %}
    </div>
</body>
</html>
"""


# INACTIVE ROUTE - uncomment @app.route to re-enable
# @app.route('/usage', methods=['GET', 'POST'])
def usage():
    results = []
    reverse_map, prod_desc = build_reverse_usage_index()

    def _dmy(s):
        try:
            dt = pd.to_datetime(s, errors='coerce')
            return dt.strftime('%d/%m/%Y') if pd.notna(dt) else ''
        except Exception:
            return ''

    if request.method == 'POST':
        given = request.form.get('materials', '').splitlines()
        given_bases = [_base_code(x) for x in given if str(x).strip()]
        # cache for per-unit materials to avoid recompute
        cache_perunit = {}
        for leaf in given_bases:
            prods = sorted(list(reverse_map.get(leaf, [])))
            # Orders that include these products
            orders = []
            if not sales_df.empty and prods:
                col_order = _find_col(sales_df, ["Παραγγελία", "Αρ. Παραγγελίας", "Παραγγελία Πώλησης", "Παραγγελία πελάτη"]) or "Παραγγελία"
                col_customer = _find_col(sales_df, ["Πελάτης", "Πελάτης/Όνομα", "Πελάτης-Όνομα"]) or "Πελάτης"
                col_item = _find_col(sales_df, ["Είδος", "Κωδικός Είδους", "Κωδ. Είδους"]) or "Είδος"
                col_open = _find_col(sales_df, ["Ανοικτή ποσότητα", "Υπόλοιπο", "Υπόλοιπο Ποσότητας", "Υπολ. Ποσότητα"]) or "Ανοικτή ποσότητα"
                col_delivery = _find_col(sales_df, ["Ημ/νία παράδοσης", "Ημερομηνία παράδοσης", "Παράδοση"]) or "Ημ/νία παράδοσης"
                col_confirmed = _find_col(sales_df, ["Επιβεβαιωμένη ημερομηνία", "Επιβεβαιωμένη", "Επιβεβαιωμ"])  # may be None
                subset = sales_df[[c for c in [col_order, col_customer, col_item, col_open, col_delivery, col_confirmed] if c in sales_df.columns]].copy()
                # Normalize base code in item
                subset['__base_item__'] = subset[col_item].astype(str).str.split().str[0].str.split('/').str[0]
                sub2 = subset[subset['__base_item__'].isin(prods)].copy()
                for _, r in sub2.iterrows():
                    orders.append({
                        'order': r.get(col_order, ''),
                        'customer': r.get(col_customer, ''),
                        'item': r.get(col_item, ''),
                        'open': pd.to_numeric(r.get(col_open, 0), errors='coerce') or 0,
                        'delivery': _dmy(r.get(col_delivery)),
                        'confirmed': _dmy(r.get(col_confirmed)) if col_confirmed and col_confirmed in sales_df.columns else ''
                    })
            # build product entries with per-unit consumption of the leaf
            prod_entries = []
            for p in prods:
                d = prod_desc.get(p, '')
                per_map = _per_unit_materials(p, cache_perunit, set())
                per_qty = None
                per_unit = ''
                if per_map:
                    total = 0.0
                    unit = None
                    for (comp_code, _comp_desc, comp_unit), q in per_map.items():
                        if _base_code(comp_code) == leaf:
                            try:
                                total += float(q or 0)
                            except Exception:
                                pass
                            if not unit:
                                unit = comp_unit
                    if total > 0:
                        per_qty = total
                        per_unit = unit or ''
                prod_entries.append({
                    'code': p,
                    'desc': d,
                    'per_unit': per_qty,
                    'unit': per_unit,
                })

            # Try to find description of the material itself from stock or recipes
            leaf_desc = ''
            try:
                # from df (recipe components)
                comp_rows = df[df['Κωδ. Αναλ.'].astype(str).str.split('/').str[0] == leaf]
                if not comp_rows.empty:
                    leaf_desc = str(comp_rows['Περιγραφή Αναλούμενου'].iloc[0])
            except Exception:
                pass

            results.append({
                'leaf': leaf,
                'desc': leaf_desc,
                'products': prod_entries,
                'orders': orders,
            })

        # If asked to download, create a flat export and return as Excel
    if request.form.get('download') == '1' and results:
            rows = []
            for r in results:
                leaf = r.get('leaf')
                leaf_desc = r.get('desc')
                for p in (r.get('products') or []):
                    rows.append({
                        'Κωδ. Αναλούμενου': leaf,
                        'Περιγραφή Αναλούμενου': leaf_desc,
                        'Κωδ. Προϊόντος': p.get('code'),
                        'Περιγραφή Προϊόντος': p.get('desc'),
                        'Ανάλωση ανά τεμάχιο': p.get('per_unit'),
                        'Μονάδα Αναλούμενου': p.get('unit'),
                    })
            out_df = pd.DataFrame(rows)
            buf = io.BytesIO()
            # Try xlsxwriter then openpyxl
            try:
                with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
                    out_df.to_excel(writer, index=False, sheet_name='Usage')
            except Exception:
                with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                    out_df.to_excel(writer, index=False, sheet_name='Usage')
            buf.seek(0)
            return send_file(
                buf,
                as_attachment=True,
                download_name=f"usage_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    return render_template_string(USAGE_TEMPLATE, results=results)

# INACTIVE ROUTE - uncomment @app.route to re-enable
# @app.route('/semis', methods=['GET','POST'])
def semis_view():
    semis = []
    details = []
    combined = []
    if request.method == 'POST':
        lines = request.form.get('items','').splitlines()
        items = []
        for ln in lines:
            parts = ln.strip().split()
            if not parts:
                continue
            code = parts[0]
            qty = 1.0
            if len(parts) > 1:
                try:
                    qty = float(parts[1])
                except Exception:
                    qty = 1.0
            items.append({'product_code': code, 'qty': qty})
        if items:
            # Υπολογισμός ημιέτοιμων
            semis, details = get_semis_multiple(items)
            # Συγκέντρωση κορυφαίων προϊόντων (ίδιος τρόπος με το /app για qty_map)
            qty_map = {}
            for it in items:
                base = _base_code(str(it.get('product_code','')))
                try:
                    q = float(it.get('qty', 0) or 0)
                except Exception:
                    q = 0
                if base:
                    qty_map[base] = qty_map.get(base, 0.0) + q

            # Δημιουργία γραμμών προϊόντων
            product_rows = []
            for base_code, q in qty_map.items():
                _rows, _bq, desc, _rev = _select_latest_recipe(base_code)
                product_rows.append({
                    'type': 'Προϊόν',
                    'code': base_code,
                    'desc': desc,
                    'qty': q,
                    'unit': 'ΤΕΜ'
                })

            # Μετατροπή ημιέτοιμων σε γραμμές με τύπο
            semi_rows = [{
                'type': 'Ημιέτοιμο',
                'code': s.get('code'),
                'desc': s.get('desc'),
                'qty': s.get('qty'),
                'unit': s.get('unit'),
            } for s in semis]

            combined = product_rows + semi_rows

    # Excel export
    if request.method == 'POST' and request.form.get('download') == '1' and (combined or semis):
        rows = []
        for r in (combined if combined else []):
            rows.append({
                'Τύπος': r.get('type'),
                'Κωδικός': r.get('code'),
                'Περιγραφή': r.get('desc'),
                'Ποσότητα': r.get('qty'),
                'Μονάδα': r.get('unit'),
            })
        out_df = pd.DataFrame(rows)
        buf = io.BytesIO()
        try:
            with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
                out_df.to_excel(writer, index=False, sheet_name='Products_Semis')
        except Exception:
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                out_df.to_excel(writer, index=False, sheet_name='Products_Semis')
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"products_semis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return render_template_string(SEMI_TEMPLATE, semis=semis, details=details, combined=combined, request=request)


USAGE_BY_STORAGE_TEMPLATE = r"""
<!doctype html>
<html>
<head>
    <meta charset=\"utf-8\">
    <title>Ανάλωση ανά Α.Χ.</title>
    <style>
        body { background:#fff; color:#000; font-family:sans-serif; margin:0; }
        .container { width:100%; padding:16px 20px; box-sizing:border-box; }
        table { width:100%; border-collapse: collapse; margin-top: 12px; }
        th, td { border:1px solid #000; padding:4px; text-align:left; }
        .num { text-align:right; font-variant-numeric: tabular-nums; }
        .btn { display:inline-block; padding:8px 12px; border:1px solid #000; background:#eee; color:#000; text-decoration:none; cursor:pointer; }
        .scroll { max-height:160px; overflow:auto; border:1px solid #ccc; padding:6px; }
    </style>
    <script>
        function parseNumberGR(s){ if(s==null) return 0; s=(''+s).trim(); if(s==='') return 0; s=s.replace(/\./g,'').replace(',', '.'); var v=parseFloat(s); return isNaN(v)?0:v; }
        function sortTableBy(tableId, colIndex, type){
            var table = document.getElementById(tableId); if(!table) return;
            var rows = Array.prototype.slice.call(table.querySelectorAll('tr')); if(rows.length===0) return;
            var header = rows.shift(); var curCol=table.getAttribute('data-sort-col'); var curDir=table.getAttribute('data-sort-dir')||'asc';
            var dir=(curCol===String(colIndex)&&curDir==='asc')?'desc':'asc'; var mod = dir==='asc'?1:-1;
            function cellText(tr, idx){ var tds=tr.getElementsByTagName('td'); return (tds[idx] && (tds[idx].innerText||tds[idx].textContent)||'').trim(); }
            rows.sort(function(a,b){
                var A=cellText(a,colIndex), B=cellText(b,colIndex);
                if(type==='number'){ var na=parseNumberGR(A), nb=parseNumberGR(B); if(na<nb) return -1*mod; if(na>nb) return 1*mod; return 0; }
                A=(A||'').toLowerCase(); B=(B||'').toLowerCase(); return A.localeCompare(B,'el',{numeric:true,sensitivity:'base'})*mod;
            });
            while(table.firstChild) table.removeChild(table.firstChild);
            table.appendChild(header); for (var i=0;i<rows.length;i++) table.appendChild(rows[i]);
            table.setAttribute('data-sort-col', String(colIndex)); table.setAttribute('data-sort-dir', dir);
            var ths = header.getElementsByTagName('th'); for (var j=0;j<ths.length;j++){ ths[j].removeAttribute('data-sorted'); ths[j].style.cursor='pointer'; }
            if (ths[colIndex]) ths[colIndex].setAttribute('data-sorted', dir);
        }
        function initSortable(tableId){ var table=document.getElementById(tableId); if(!table) return; var header=table.querySelector('tr'); if(!header) return; var ths=header.getElementsByTagName('th'); for (var i=0;i<ths.length;i++){ (function(idx){ var th=ths[idx]; var t=th.getAttribute('data-sort')||'text'; if(t==='none') return; th.style.cursor='pointer'; th.onclick=function(){ sortTableBy(tableId, idx, t); }; })(i);} }
        window.addEventListener('load', function(){ initSortable('ax_usage_table'); });
        function downloadAxExcel(){
            var f = document.getElementById('ax_form'); if(!f) return;
            var inp = document.createElement('input'); inp.type='hidden'; inp.name='download'; inp.value='1'; f.appendChild(inp);
            f.submit(); setTimeout(function(){ if(inp && inp.parentNode) inp.parentNode.removeChild(inp); }, 0);
        }
    </script>
</head>
<body>
    <div class=\"container\">
        <h2>Ανάλωση ανά Α.Χ.</h2>
    <form method=\"post\" id=\"ax_form\">
            <div>
                <strong>Επιλογή Α.Χ.:</strong>
                {% if storage_list %}
                <div class=\"scroll\">
                    {% for ax in storage_list %}
                    <label style=\"display:inline-block; min-width:140px;\">
                        <input type=\"checkbox\" name=\"ax\" value=\"{{ ax }}\" {% if ax in selected_ax %}checked{% endif %}>
                        {{ ax }}
                    </label>
                    {% endfor %}
                </div>
                {% else %}
                <div>Δεν βρέθηκαν αποθηκευτικοί χώροι (Α.Χ.).</div>
                {% endif %}
            </div>
            <div style=\"margin-top:8px;\">
                <button class=\"btn\" type=\"submit\">Ανάλυση</button>
                <a class=\"btn\" href=\"/\">Αρχική</a>
                {% if results %}
                <button class=\"btn\" type=\"button\" onclick=\"downloadAxExcel()\">Κατέβασμα σε Excel</button>
                {% endif %}
            </div>
        </form>

    {% if results %}
        <table id=\"ax_usage_table\">
            <tr>
                <th data-sort=\"text\">Α.Χ.</th>
                <th data-sort=\"text\">Κωδ. Υλικού</th>
                <th data-sort=\"text\">Περιγραφή</th>
                <th class=\"num\" data-sort=\"number\">Διαθέσιμο στην Α.Χ.</th>
                <th data-sort=\"text\">Αναλώνεται σε προϊόντα</th>
            </tr>
            {% for r in results %}
            <tr>
                <td>{{ r.ax }}</td>
                <td>{{ r.code }}</td>
                <td>{{ r.desc }}</td>
                <td class=\"num\">{{ r.qty|gr(2) }}</td>
                <td>{{ r.used_in }}</td>
            </tr>
            {% endfor %}
        </table>
        {% elif selected_ax %}
            <p>Δεν βρέθηκαν υλικά στα επιλεγμένα Α.Χ.</p>
        {% endif %}
    </div>
</body>
</html>
"""


# INACTIVE ROUTE - uncomment @app.route to re-enable
# @app.route('/usage-by-storage', methods=['GET', 'POST'])
def usage_by_storage():
    storage_list = get_storage_list(stock_df)
    selected_ax = request.form.getlist('ax') if request.method == 'POST' else []
    results = []
    if request.method == 'POST' and selected_ax:
        # Build stock index and reverse usage index
        stock_index = build_stock_index(stock_df)
        reverse_map, prod_desc = build_reverse_usage_index()
        # Reverse lookup helper for material description (from recipes components)
        def get_leaf_desc(code_base: str):
            try:
                comp_rows = df[df['Κωδ. Αναλ.'].astype(str).str.split('/').str[0] == code_base]
                if not comp_rows.empty:
                    return str(comp_rows['Περιγραφή Αναλούμενου'].iloc[0])
            except Exception:
                pass
            return ''
        # Iterate materials and pick quantities per selected A.X.
        for code, info in (stock_index or {}).items():
            by = (info.get('by_storage', {}) or {})
            for ax in selected_ax:
                qty = float(by.get(ax, 0) or 0)
                if qty <= 0:
                    continue
                base_code = _base_code(code)
                used_products = []
                for p in sorted(list(reverse_map.get(base_code, []))):
                    d = prod_desc.get(p, '')
                    used_products.append(f"{p}{(' ('+d+')') if d else ''}")
                results.append({
                    'ax': ax,
                    'code': base_code,
                    'desc': get_leaf_desc(base_code),
                    'qty': qty,
                    'used_in': ", ".join(used_products)
                })
        # Optional: sort results by AX then code initially
        results.sort(key=lambda x: (str(x['ax']), str(x['code'])))
        # If download requested, build and return Excel
        if request.form.get('download') == '1' and results:
            out_df = pd.DataFrame(results)
            # Reorder columns nicely
            cols = ['ax','code','desc','qty','used_in']
            out_df = out_df[[c for c in cols if c in out_df.columns]]
            out_df = out_df.rename(columns={'ax':'Α.Χ.','code':'Κωδ. Υλικού','desc':'Περιγραφή','qty':'Διαθέσιμο στην Α.Χ.','used_in':'Αναλώνεται σε προϊόντα'})
            buf = io.BytesIO()
            try:
                with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
                    out_df.to_excel(writer, index=False, sheet_name='UsageByAX')
            except Exception:
                with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                    out_df.to_excel(writer, index=False, sheet_name='UsageByAX')
            buf.seek(0)
            return send_file(
                buf,
                as_attachment=True,
                download_name=f"usage_by_ax_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    return render_template_string(USAGE_BY_STORAGE_TEMPLATE, storage_list=storage_list, selected_ax=selected_ax, results=results)


# --------------------------- Production Plan ---------------------------
# Λίστα γραμμών παραγωγής
PRODUCTION_LINES = [
    "DISPENSING AREA - PR.0.38",
    "GRANULATION No 1 - HIGH SHEAR MIXER - PR.1.09",
    "GRANULATION No 1 - FLUID BED DRYER - PR.1.10",
    "GRANULATION No 2/LOW SHEAR MIXER - PR.0.43",
    "GRANULATION No 3 - MIXER 1200L - PR.0.44",
    "TABLETING No 1 - PR.1.08",
    "TABLETING No 2 - fette - PR.1.08",
    "COATING - PR.1.04",
    "CAPSULATION No 1/ZANASI 12F - PR.1.06",
    "CAPSULATION No 2/ZANASI 40F - PR.1.02",
    "LIQUIDS PREPARATION/FILLING No 1 (OMAS) - PR.1.15",
    "LIQUIDS PREPARATION/FILLING No 2 (PMS)- PR.1.16",
    "LIQUIDS PREPARATION No 3 - PR.0.29",
    "LIQUIDS FILLING No 3 (TECNOMACO) - PR.0.30",
    "SACHET FILLING - PR.1.03",
    "BLISTER PACKAGING No 1 - (Noack DPN760 Alu/Alu)-PA.1.18",
    "BLISTER PACKAGING No 4 - PA.1.19 (Noack DPN 760)",
    "BLISTER PACKAGING No 2 - (UHLMANN B1240) - PA.1.20",
    "SECONDARY PACKAGING No 2 -  (UHLMANN C130) - PA.1.21",
    "SECONDARY PACKAGING No 1 - (Rotary Baloglou) - PA.1.22",
    "BOTTLE PACKAGING No 3 - (Pharma Packaging) - PA.0.31",
    "SECONDARY PACKAGING No 3 - (PROMATIC BIPAK) - PA.0.32",
    "TERTIARY PACKAGING No 1  (SERIALIZATION EQUIPMENT) -PA.0.42",
]

def list_fg_and_bulk_from_recipes():
    """Επιστρέφει (fg_list, bulk_list) από το receipes.xlsx.
    FG: μοτίβο xx1-xx-xx, BULK: xx2-xx-xx (βάση κωδικού).
    """
    if df.empty:
        return [], []
    work = df.copy()
    bases = []
    if 'Κωδικός Είδους Συνταγής' in work.columns:
        try:
            bases = work['Κωδικός Είδους Συνταγής'].astype(str).str.split('/').str[0].dropna().unique().tolist()
        except Exception:
            bases = []
    elif 'Κωδικός' in work.columns:
        try:
            bases = work['Κωδικός'].astype(str).str.split('/').str[0].dropna().unique().tolist()
        except Exception:
            bases = []
    fg = sorted([b for b in bases if re.match(r'^\d{2}1-\d{2}-\d{2}$', str(b))])
    bulk = sorted([b for b in bases if re.match(r'^\d{2}2-\d{2}-\d{2}$', str(b))])
    return fg, bulk

def per_unit_of_material_in_product(leaf_base: str, product_base: str):
    """Επιστρέφει (ποσότητα ανά τεμάχιο, μονάδα) ενός leaf στο προϊόν.
    Χρησιμοποιεί την _per_unit_materials για πλήρη ανάπτυξη.
    """
    per_map = _per_unit_materials(product_base, {}, set())
    total = 0.0
    unit = ''
    for (comp_code, _desc, comp_unit), q in per_map.items():
        if _base_code(comp_code) == leaf_base:
            try:
                total += float(q or 0)
            except Exception:
                pass
            if not unit:
                unit = comp_unit or ''
    return total, unit

PLAN_TEMPLATE = """
<!doctype html>
<html>
<head>
    <meta charset=\"utf-8\">
    <title>Production Plan</title>
    <style>
        body { background:#fff; color:#000; font-family:sans-serif; margin:0; }
        .container { width:100%; padding:16px 20px; box-sizing:border-box; }
        table { width:100%; border-collapse: collapse; margin-top: 12px; }
        th, td { border:1px solid #000; padding:4px; text-align:left; vertical-align: top; }
        .num { text-align:right; font-variant-numeric: tabular-nums; }
        .grid { display:grid; grid-template-columns: repeat(4, minmax(220px, 1fr)); gap:12px; }
        .scroll { max-height:180px; overflow:auto; border:1px solid #ccc; padding:6px; }
        .btn { display:inline-block; padding:8px 12px; border:1px solid #000; background:#eee; color:#000; text-decoration:none; cursor:pointer; }
        label { display:block; }
        textarea { width:100%; height:90px; }
    </style>
    <script>
        function debounce(fn, delay){ var t; return function(){ var ctx=this, args=arguments; clearTimeout(t); t=setTimeout(function(){ fn.apply(ctx,args); }, delay||300); } }
        function qs(sel, root){ return (root||document).querySelector(sel); }
        function cellInputs(li, dj){
            var c = qs('[name="sched_'+li+'_'+dj+'_code"]');
            var q = qs('[name="sched_'+li+'_'+dj+'_qty"]');
            return {code:c, qty:q};
        }
        function updateDailyTotals(){
            try{
                var perShift = parseFloat('{{ people_per_shift }}') || 0;
                var shifts = parseInt('{{ shifts }}') || 0;
                var selected = new Set([
                    {% for ln in selected_lines %}'{{ ln|replace("'","\'") }}',{% endfor %}
                ]);
                var days = {{ dates|length if dates else 0 }};
                for(var dj=0; dj<days; dj++){
                    var count = 0;
                    for(var li=0; li<{{ all_lines|length if all_lines else 0 }}; li++){
                        var line = {{ all_lines|tojson }}[li];
                        if(!selected.has(line)) continue;
                        var ins = cellInputs(li, dj);
                        var code = ins.code ? (ins.code.value||'').trim() : '';
                        if(code){ count += 1; }
                    }
                    var val = count * perShift * shifts;
                    var th = document.getElementById('people_day_'+dj);
                    if(th){ th.textContent = (val || 0).toLocaleString('el-GR', {minimumFractionDigits:2, maximumFractionDigits:2}); }
                }
            }catch(e){ /* ignore */ }
        }
        var saveCellDebounced = debounce(function(li, dj){
            var line = {{ all_lines|tojson if all_lines else '[]' }}[li];
            var d = {{ dates|tojson if dates else '[]' }}[dj];
            if(!line || !d) return;
            var ins = cellInputs(li, dj);
            var code = ins.code ? ins.code.value.trim() : '';
            var qty = ins.qty ? ins.qty.value.trim() : '';
            var st = document.getElementById('cell_status_'+li+'_'+dj);
            if(st){ st.textContent = '⏳'; st.style.color = '#555'; }
            fetch('/plan/save-cell', {
                method: 'POST', headers: {'Content-Type':'application/json'},
                body: JSON.stringify({ line: line, date: d.iso, code: code, qty: qty })
            }).then(function(r){ return r.json(); }).then(function(resp){
                if(st){ st.textContent = resp.ok ? '✓' : '×'; st.style.color = resp.ok ? 'green' : 'red'; }
                updateDailyTotals();
            }).catch(function(){ if(st){ st.textContent = '×'; st.style.color = 'red'; } });
        }, 400);
        function onCellChange(li, dj){ saveCellDebounced(li, dj); }
        window.addEventListener('load', function(){ updateDailyTotals(); });
    </script>
</head>
<body>
<div class=\"container\">
    <h2>Production Plan</h2>
    <form method=\"post\">
        <div class=\"grid\">
            <div>
                <strong>Πλάνο (κωδικός ποσότητα)</strong>
                <small>π.χ. 121-00-01 10000 ή 122-00-01 5000</small>
                <textarea name=\"plan_items\">{{ plan_items }}</textarea>
                <div style=\"margin-top:6px;\"><small>Έτοιμα (xx1): {{ fg|length }} — Bulk (xx2): {{ bulk|length }}</small></div>
            </div>
            <div>
                <strong>Γραμμές παραγωγής</strong>
                <div class=\"scroll\">
                    {% for line in lines %}
                        <label><input type=\"checkbox\" name=\"lines\" value=\"{{ line }}\" {% if line in selected_lines %}checked{% endif %}> {{ line }}</label>
                    {% endfor %}
                </div>
            </div>
            <div>
                <strong>Βάρδιες</strong>
                <label>Αριθμός βαρδιών:
                    <select name=\"shifts\">{% for s in [1,2,3] %}<option value=\"{{ s }}\" {% if s==shifts %}selected{% endif %}>{{ s }}</option>{% endfor %}</select>
                </label>
                <label>Ώρες/βάρδια: <input type=\"number\" name=\"hours_per_shift\" step=\"0.5\" value=\"{{ hours_per_shift }}\"></label>
                <label>Άτομα/βάρδια: <input type=\"number\" name=\"people_per_shift\" step=\"1\" value=\"{{ people_per_shift }}\"></label>
            </div>
            <div>
                <strong>Διάστημα</strong>
                <label>Από: <input type=\"date\" name=\"start_date\" value=\"{{ start_date }}\"></label>
                <label>Έως: <input type=\"date\" name=\"end_date\" value=\"{{ end_date }}\"></label>
            </div>
        </div>
        <div style=\"margin-top:12px;\">
            <button class=\"btn\" type=\"submit\">Υπολογισμός</button>
            <a class=\"btn\" href=\"/\">Αρχική</a>
        </div>
    </form>

    {% if show_grid and dates and all_lines %}
    <hr>
    <h3>Πλάνο ανά γραμμή και ημέρα</h3>
    <form method="post">
        <!-- Preserve parameters -->
        <input type="hidden" name="start_date" value="{{ start_date }}">
        <input type="hidden" name="end_date" value="{{ end_date }}">
        <input type="hidden" name="shifts" value="{{ shifts }}">
        <input type="hidden" name="hours_per_shift" value="{{ hours_per_shift }}">
        <input type="hidden" name="people_per_shift" value="{{ people_per_shift }}">
        <input type="hidden" name="plan_items" value="{{ plan_items }}">
        {% for line in selected_lines %}
            <input type="hidden" name="lines" value="{{ line }}">
        {% endfor %}

        <!-- Choices for κωδικοί: FG και BULK -->
        <datalist id="codes_list">
            {% for c in fg %}<option value="{{ c }}"></option>{% endfor %}
            {% for c in bulk %}<option value="{{ c }}"></option>{% endfor %}
        </datalist>

        <table>
            <tr>
                <th>Γραμμή</th>
                {% for d in dates %}
                    <th>{{ d.label }}</th>
                {% endfor %}
            </tr>
            {% for line in all_lines %}
            {% set li = loop.index0 %}
            <tr>
                <td>{{ line }}</td>
                {% for d in dates %}
                {% set dj = loop.index0 %}
                {% set cell = (schedule_map.get(line, {}).get(d.iso, {})) %}
                {% set disabled = (line not in selected_lines) %}
                <td style="vertical-align:top; {% if disabled %}background:#f7f7f7;{% endif %}">
                    <div>
                        <input name="sched_{{ li }}_{{ dj }}_code" list="codes_list" placeholder="κωδ." value="{{ cell.get('code','') }}" {% if disabled %}disabled{% endif %} style="width:120px;" oninput="onCellChange({{ li }}, {{ dj }})">
                    </div>
                    <div>
                        <input name="sched_{{ li }}_{{ dj }}_qty" type="number" step="0.01" placeholder="ποσ." value="{{ cell.get('qty','') }}" {% if disabled %}disabled{% endif %} style="width:100px;" oninput="onCellChange({{ li }}, {{ dj }})">
                        <span id="cell_status_{{ li }}_{{ dj }}" style="margin-left:6px; font-size:12px; color:#555;"></span>
                    </div>
                </td>
                {% endfor %}
            </tr>
            {% endfor %}
            <tr>
                <th>Σύνολο ατόμων/ημέρα</th>
                {% for p in daily_people %}
                    <th class="num" id="people_day_{{ loop.index0 }}">{{ p|gr(2) }}</th>
                {% endfor %}
            </tr>
        </table>
        <div style="margin-top:10px;">
            <button class="btn" type="submit">Υπολογισμός/Αποθήκευση πλάνου</button>
        </div>
    </form>
    {% endif %}

    {% if summary %}
    <h3>Σύνοψη δυναμικότητας</h3>
    <table>
        <tr><th>Επιλεγμένες γραμμές</th><td>{{ summary.lines_count }}</td></tr>
        <tr><th>Βάρδιες/ημέρα</th><td>{{ summary.shifts }}</td></tr>
        <tr><th>Ώρες/βάρδια</th><td class=\"num\">{{ summary.hours_per_shift|gr(2) }}</td></tr>
        <tr><th>Άτομα/βάρδια</th><td class=\"num\">{{ summary.people_per_shift|gr(2) }}</td></tr>
        <tr><th>Σύνολο ατόμων ανά ημέρα</th><td class=\"num\">{{ summary.people_per_day|gr(2) }}</td></tr>
        <tr><th>Ημέρες</th><td class=\"num\">{{ summary.days }}</td></tr>
        <tr><th>Συνολικά άτομα (μέρες x άτομα/ημέρα)</th><td class=\"num\">{{ summary.people_total|gr(2) }}</td></tr>
    </table>

    {% if plan_results %}
    <h3>Ανάλυση πλάνου</h3>
    <table>
        <tr>
            <th>Κωδικός</th>
            <th>Περιγραφή</th>
            <th>Τύπος</th>
            <th class=\"num\">Ποσότητα</th>
            <th>Συσχέτιση</th>
        </tr>
        {% for r in plan_results %}
        <tr>
            <td>{{ r.code }}</td>
            <td>{{ r.desc }}</td>
            <td>{{ r.kind }}</td>
            <td class=\"num\">{{ r.qty|gr(2) }}</td>
            <td>
                {% if r.kind == 'BULK' and r.links %}
                    <div><strong>Σχετικά έτοιμα:</strong></div>
                    <ul style=\"margin:6px 0 0 18px; padding:0;\">
                        {% for l in r.links %}
                            <li>{{ l.product }}{% if l.desc %} — {{ l.desc }}{% endif %} — ανά τεμάχιο: {{ l.per_unit|gr(2) }} {{ l.unit }}{% if l.possible is not none %} — εκτ. κάλυψη: {{ l.possible|gr(2) }} τεμ.{% endif %}</li>
                        {% endfor %}
                    </ul>
                {% else %}
                    —
                {% endif %}
            </td>
        </tr>
        {% endfor %}
    </table>
    {% endif %}
    {% endif %}
</div>
</body>
</html>
"""

# INACTIVE ROUTE - uncomment @app.route to re-enable
# @app.route('/plan', methods=['GET','POST'])
def plan():
    fg, bulk = list_fg_and_bulk_from_recipes()

    # Defaults
    plan_items_text = request.form.get('plan_items', '') if request.method == 'POST' else ''
    selected_lines = request.form.getlist('lines') if request.method == 'POST' else []
    try:
        shifts = int(request.form.get('shifts', 1))
    except Exception:
        shifts = 1
    try:
        hours_per_shift = float(request.form.get('hours_per_shift', 8))
    except Exception:
        hours_per_shift = 8.0
    try:
        people_per_shift = float(request.form.get('people_per_shift', 4))
    except Exception:
        people_per_shift = 4.0
    start_date = request.form.get('start_date', '')
    end_date = request.form.get('end_date', '')

    summary = None
    plan_results = []
    # Grid: dates and schedule values
    dates = []  # list of dicts: {iso, label}
    schedule_map = {}
    daily_people = []
    show_grid = False

    if request.method == 'POST':
        # Ημέρες στο διάστημα (συμπεριλαμβανόμενες)
        days = 0
        try:
            if start_date and end_date:
                s = pd.to_datetime(start_date, errors='coerce')
                e = pd.to_datetime(end_date, errors='coerce')
                if pd.notna(s) and pd.notna(e) and e >= s:
                    # Build dates list and labels (dd/mm (Δε))
                    cur = s.normalize()
                    greek = ['Δε','Τρ','Τε','Πε','Πα','Σα','Κυ']
                    while cur <= e:
                        d_iso = cur.date().isoformat()
                        label = f"{cur.strftime('%d/%m')} ({greek[int(cur.dayofweek)]})"
                        dates.append({'iso': d_iso, 'label': label})
                        cur = cur + pd.Timedelta(days=1)
                    days = len(dates)
                    show_grid = True if selected_lines else False
        except Exception:
            days = 0
        lines_count = len(selected_lines)
        people_per_day = people_per_shift * shifts * max(lines_count, 0)
        people_total = people_per_day * max(days, 0)
        summary = {
            'lines_count': lines_count,
            'shifts': shifts,
            'hours_per_shift': hours_per_shift,
            'people_per_shift': people_per_shift,
            'people_per_day': people_per_day,
            'days': days,
            'people_total': people_total,
        }

        # Ανάλυση πλάνου: κάθε γραμμή "code qty"
        entries = []
        for ln in plan_items_text.splitlines():
            parts = ln.strip().split()
            if not parts:
                continue
            code = _base_code(parts[0])
            qty = 0.0
            if len(parts) >= 2:
                try:
                    qty = float(parts[1])
                except Exception:
                    qty = 0.0
            entries.append((code, qty))

        # Περιγραφή προϊόντος από συνταγές
        def _desc_of(code_base: str) -> str:
            rows, _, desc, _ = _select_latest_recipe(code_base)
            return desc or ''

        reverse_map, prod_desc = build_reverse_usage_index()
        for code, qty in entries:
            kind = 'FG' if re.match(r'^\d{2}1-\d{2}-\d{2}$', code or '') else ('BULK' if re.match(r'^\d{2}2-\d{2}-\d{2}$', code or '') else 'OTHER')
            desc = _desc_of(code)
            item = {'code': code, 'desc': desc, 'kind': kind, 'qty': qty, 'links': []}
            if kind == 'BULK':
                products = sorted(list(reverse_map.get(code, [])))
                links = []
                for p in products:
                    per_u, unit = per_unit_of_material_in_product(code, p)
                    if per_u and per_u > 0:
                        possible = (qty / per_u) if qty and qty > 0 else None
                    else:
                        possible = None
                    links.append({'product': p, 'desc': prod_desc.get(p, ''), 'per_unit': per_u or 0.0, 'unit': unit or '', 'possible': possible})
                item['links'] = links
            plan_results.append(item)

        # Prefill schedule_map from PLAN_DATA for current date range and all lines
        if dates and not schedule_map:
            for ln in PRODUCTION_LINES:
                for d in dates:
                    key = (ln, d['iso'])
                    if key in PLAN_DATA['cells']:
                        schedule_map.setdefault(ln, {})[d['iso']] = PLAN_DATA['cells'][key].copy()

        # Parse schedule grid inputs (non-AJAX) and compute daily people totals
        if dates:
            # map line text to index for input names
            line_to_index = {ln: idx for idx, ln in enumerate(PRODUCTION_LINES)}
            per_day_count = [0 for _ in dates]
            for ln in PRODUCTION_LINES:
                li = line_to_index.get(ln, None)
                if li is None:
                    continue
                row_map = {}
                for dj, d in enumerate(dates):
                    code_key = f"sched_{li}_{dj}_code"
                    qty_key = f"sched_{li}_{dj}_qty"
                    code_val = request.form.get(code_key, '').strip()
                    qty_val = request.form.get(qty_key, '').strip()
                    # If no form value, keep existing from PLAN_DATA prefill
                    if not code_val and not qty_val:
                        cell = schedule_map.get(ln, {}).get(d['iso'], None)
                        if cell:
                            code_val = str(cell.get('code','')).strip()
                            qty_val = str(cell.get('qty','')).strip()
                    if code_val or qty_val:
                        try:
                            qty_num = float(qty_val) if qty_val else ''
                        except Exception:
                            qty_num = qty_val
                        row_map[d['iso']] = {'code': code_val, 'qty': qty_num}
                        # Persist in memory store
                        PLAN_DATA['cells'][(ln, d['iso'])] = {'code': code_val, 'qty': qty_num}
                        if ln in selected_lines and code_val:
                            per_day_count[dj] += 1
                if row_map:
                    schedule_map[ln] = row_map
            # people per day: scheduled lines that day * people_per_shift * shifts
            try:
                ppl_per_shift = float(people_per_shift or 0)
            except Exception:
                ppl_per_shift = 0.0
            try:
                sh = int(shifts or 0)
            except Exception:
                sh = 0
            daily_people = [p * ppl_per_shift * sh for p in per_day_count]

    return render_template_string(PLAN_TEMPLATE,
        plan_items=plan_items_text,
        lines=PRODUCTION_LINES,
        all_lines=PRODUCTION_LINES,
        selected_lines=selected_lines,
        shifts=shifts,
        hours_per_shift=hours_per_shift,
        people_per_shift=people_per_shift,
        start_date=start_date,
        end_date=end_date,
        fg=list(fg), bulk=list(bulk),
        summary=summary,
        plan_results=plan_results,
        show_grid=show_grid,
        dates=dates,
        schedule_map=schedule_map,
        daily_people=daily_people)

# INACTIVE ROUTE - uncomment @app.post to re-enable
# @app.post('/plan/save-cell')
def plan_save_cell():
    try:
        data = request.get_json(silent=True) or {}
        line = str(data.get('line','')).strip()
        date_iso = str(data.get('date','')).strip()
        code = str(data.get('code','')).strip()
        qty_raw = str(data.get('qty','')).strip()
        if not line or line not in PRODUCTION_LINES:
            return {'ok': False, 'error': 'invalid line'}, 400
        # basic date validation
        try:
            pd.to_datetime(date_iso, errors='raise')
        except Exception:
            return {'ok': False, 'error': 'invalid date'}, 400
        try:
            qty = float(qty_raw) if qty_raw != '' else ''
        except Exception:
            qty = qty_raw
        PLAN_DATA['cells'][(line, date_iso)] = {'code': code, 'qty': qty}
        return {'ok': True}
    except Exception as e:
        return {'ok': False, 'error': str(e)}, 500

app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

if __name__ == '__main__':
    app.run(debug=True)
